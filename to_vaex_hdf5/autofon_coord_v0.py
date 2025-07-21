# !/usr/bin/env python3
# coding:utf-8
"""
  Author:  Andrey Korzh <ao.korzh@gmail.com>
  Purpose: Load coordinates from autofon (http://www.autofon.ru/autofon/item/seplus) GPS trackers to GPX files and HDF5 pandas store
  Created: 08.04.2021
  Modified: 29.06.2021
"""
import sys
import logging
from pathlib import Path
from shutil import copyfile
from typing import Any, Callable, Dict, Iterator, Mapping, MutableMapping, Optional, List, Tuple, Union
from datetime import datetime, timedelta, timezone
from itertools import zip_longest
from dataclasses import dataclass, field
import hydra
import numpy as np
import tables
import pandas as pd
from pandas.tseries.frequencies import to_offset
from tables.exceptions import HDF5ExtError
import re
import requests
# import gc
from time import sleep
from tabulate import tabulate
from gpxpy.gpx import GPX
# import pyproj   # from geopy import Point, distance
from h5toGpx import save_to_gpx  # gpx_track_create
# from to_pandas_hdf5.h5_dask_pandas import h5.append, filter_global_minmax, filter_local
from to_pandas_hdf5.h5toh5 import h5.unzip_if_need, h5.append_log, h5.remove_tables, h5.dispenser_and_names_gen

import cfg_dataclasses
from utils2init import LoggingStyleAdapter, FakeContextIfOpen, set_field_if_no, Ex_nothing_done, call_with_valid_kwargs, ExitStatus, GetMutex

# from csv2h5_vaex import argparser_files, with_prog_config
from to_pandas_hdf5.h5toh5 import h5.move_tables, h5.index_sort, h5.out_init, h5.replace_bad_db  #, h5.rem_last_rows
from to_pandas_hdf5.gpx2h5 import h5_sort_filt_append  # df_rename_cols,
from gps_tracker.mail_parse import spot_tracker_data_from_mbox, spot_from_gmail
# from inclinometer.incl_h5clc import dekart2polar_df_uv

lf = LoggingStyleAdapter(logging.getLogger(__name__))
cfg = None
tables2mid = {
    'tr0': 221910,
    'tr2': 221912,
    'sp2': 2575092, # ESN
    'sp3': 3124620,
    'sp4': 3125300,
    'sp5': 3125411,
    'sp6': 3126104
    }
mid2tables = {v: k for k, v in tables2mid.items()}

# cfg = {
#     'in': {
#         'tables': ['tr0'],  # '*',   # what to process
#         'time_intervals': {
#             'tr0': [pd.Timestamp('2021-04-08T12:00:00'), pd.Timestamp('2021-04-08T12:00:00')]},
#         'dt_from_utc': timedelta(hours=3),
#         # use already loaded coordinates instead of request:
#         'path_raw_local': {
#             'tr0': Path('d:\Work') /
#                     r' координаты адреса выходов на связь 09-04-2021 16-46-52 09-04-2021 16-46-52.xlsx'
#             },
#         'anchor': '[44.56905, 37.97308]',
#         },
#     'out': {
#         'path': Path(r'd:\WorkData\BlackSea\210408_trackers\210408trackers.h5'),
#         },
#     'process':
#         {
#             'simplify_tracks_error_m': 0,
#             'dt_per_file': timedelta(days=356),
#             'b_missed_coord_to_zeros': False
#             }
#     }


def save2gpx(nav_df: pd.DataFrame,
             track_name: str,
             path: Path,
             process: Dict[str, Any],
             gpx: Optional[GPX] = None,
             dt_from_utc: Optional[timedelta] = None) -> GPX:
    """
    Saves track and point process['anchor_coord'] to the ``path / f"{nav_df.index[0]:%y%m%d_%H%M}{track_name}.gpx"``
    :param nav_df: DataFrame
    :param track_name:
    :param path: gpx path
    :param process: fields:
        anchor_coord: List[float], [Lat, Lon] degrees, not used if anchor_coord_time_dict specified
        anchor_coord_time_dict: Dict[time_str, List[float]], {time_str: [Lat, Lon]} degrees. Only last item used
        anchor_depth: float, m
    :param gpx: gpxpy.gpx
    :param dt_from_utc:
    :return: updated gpx
    """
    nav_df.index.name = 'Time'
    gpx = save_to_gpx(
        nav_df if dt_from_utc is None else nav_df.tz_convert(timezone(dt_from_utc)),
        None,
        gpx_obj_namef=track_name, cfg_proc=process, gpx=gpx)

    if any(process['anchor_coord_time_dict']):
        key_last = list(process['anchor_coord_time_dict'].keys())[-1]
        lat_lon = process['anchor_coord_time_dict'][key_last]
        tim = [pd.Timestamp(key_last, tz='utc')]
    else:
        lat_lon = process['anchor_coord']
        tim = nav_df.index[[0]]

    return save_to_gpx(
        pd.DataFrame(
            {
                'Lat': lat_lon[0],
                'Lon': lat_lon[1],
                'DepEcho': process['anchor_depth'],
                'itbl': 0
             },
            index=tim
                     ),
        path.with_name(f"{nav_df.index[0]:%y%m%d_%H%M}{track_name}"),
        waypoint_symbf='Anchor',
        gpx_obj_namef='Anchor', cfg_proc=process, gpx=gpx
     )


def prepare_loading_xlsx_links_by_pandas():
    """xlsx hyperlink support to pandas, modified original https://github.com/pandas-dev/pandas/issues/13439"""

    if prepare_loading_xlsx_links_by_pandas.is_done:
        return()

    from pandas.io.excel._openpyxl import OpenpyxlReader
    from pandas._typing import FilePathOrBuffer, Scalar
    from openpyxl.cell.cell import TYPE_BOOL, TYPE_ERROR, TYPE_NUMERIC, TIME_TYPES

    def _convert_cell(self, cell, convert_float: bool) -> Scalar:

        # here we adding this hyperlink support:
        if cell.hyperlink and cell.hyperlink.target:
            return cell.hyperlink.target
            # just for example, you able to return both value and hyperlink,
            # comment return above and uncomment return below
            # btw this may hurt you on parsing values, if symbols "|||" in value or hyperlink.
            # return f'{cell.value}|||{cell.hyperlink.target}'
        # here starts original code, except for "if" became "elif"
        else:
            cell_type = cell.data_type
            if cell_type in TIME_TYPES:
                return cell.value
            elif cell_type == TYPE_ERROR:
                return np.nan
            elif cell_type == TYPE_BOOL:
                return bool(cell.value)
            elif cell.value is None:
                return ""  # compat with xlrd
            elif cell_type == TYPE_NUMERIC:
                # GH5394
                if convert_float:
                    val = int(cell.value)
                    if val == cell.value:
                        return val
                else:
                    return float(cell.value)

        return cell.value

    def load_workbook(self, filepath_or_buffer: FilePathOrBuffer):
        from openpyxl import load_workbook
        # had to change read_only to False:
        return load_workbook(filepath_or_buffer, read_only=False, data_only=True, keep_links=False)

    OpenpyxlReader._convert_cell = _convert_cell
    OpenpyxlReader.load_workbook = load_workbook
    prepare_loading_xlsx_links_by_pandas.is_done = True


prepare_loading_xlsx_links_by_pandas.is_done = False


def autofon_df_from_dict(g, dt_from_utc: timedelta):
    """

    :param g:
    :return:

    g can contain:
     {'id': 2415919104,
      'fmt': 'CAN 0: {0} [Скорость: {2} км/ч] | CAN 1: {1} [Расход топлива: {3} л]', 'lvl': 4},  # no data
     {'id': 2550136832,
      'fmt': 'Температура 1: 0,0 С 2: 0,0 С 3: 0,0 С 4: 0,0 С', 'lvl': 5},                       # no data
     {'id': 2533359616,
      'fmt': 'Общий пробег по данным GPS/ГЛОНАСС: {0} м', 'lvl': 4},                             # no data
     {'id': 2524971008,
     'fmt': 'Курс: {0}, высота: {1}, ускорение: {2}', 'lvl': 4},                                 # rough Курс data
     {'id': 2499805184,
      'fmt': 'V1: {0} В | V2: {1} В | Vвнеш: {2} В | Vбат: {3} В', 'lvl': 4},                    # rough data
     {'id': 2155872256,
     'fmt': 'Широта: {0} Долгота: {1} Скорость: {2}', 'lvl': 4},                                 # used

     {'id': 2491416576,
      'fmt': 'Уровень сигнала GSM: {0}, HDOP: {1}, спутников GPS: {2}, спутников ГЛОНАСС: {3}, температура терминала: {4} °С',
      'lvl': 4},

     {'id': 2214789120,
      'fmt': 'Конфигурация: Параметры периода опроса координат GPS',
      'lvl': 4},
     {'id': 2214854656, 'fmt': 'Конфигурация: Первая группа параметров', 'lvl': 4},
     {'id': 2214920192, 'fmt': 'Конфигурация: Вторая группа параметров', 'lvl': 4},
     {'id': 2214985728, 'fmt': 'Конфигурация: Третья группа параметров', 'lvl': 4},
     {'id': 2215051264,
      'fmt': 'Конфигурация: Четвертая группа параметров',
      'lvl': 4},
     {'id': 2215116800, 'fmt': 'Конфигурация: Пятая группа параметров', 'lvl': 4},
     {'id': 3187671040, 'fmt': '{0}', 'lvl': 4} - на самом деле параметры
     ...
     We selected:
    id_lat_lon_speed = 2155872256
    id_lgsm_hdop_ngps_nglonass = 2491416576
    HDOP: Dilution of precision (DOP): Ideal if <= 1
    """
    id_coords = 2155872256
    id_config = 3187671040

    # used parameters
    prm = {
        id_coords: {'cols': 'Lat Lon Speed'.split(),
                    'types': np.float32},
        2491416576: {'cols': 'LGSM HDOP n_GPS n_GLONASS Temp'.split(),
                     'types': {'LGSM': np.int8, 'HDOP': np.float16, 'n_GPS': np.int8, 'Temp': np.int8},
                     'drop': ['n_GLONASS'],
                     'comma_to_dot': ['HDOP']},
        2524971008: {'cols': 'Course Height Acceleration'.split(),
                     'types': np.int8,
                     'drop': ['Height', 'Acceleration'],
                     'comma_to_dot': ['Course']},
        id_config: {'cols': 'device_settings'}
        }

    # id_lat_lon_speed = 2155872256
    # id_lgsm_hdop_ngps_nglonass = 2491416576
    # id_course_heigt_acc = 2524971008

    # cols = (['Lat', 'Lon', 'Speed'], ['LGSM', 'HDOP', 'n_GPS', 'n_GLONASS', 'Temp'])
    # cols_drop = ([], ['HDOP', 'n_GLONASS'])  # strange values for HDOP (why comma separated?), n_GLONASS always zero?
    # cols_types = (np.float32, np.int8)
    # p_lat_lon_speed = []
    # p_lgsm_hdop_ngps_nglonass = []
    p = {p_id: [] for p_id in prm}
    p_detected = {p_id: False for p_id in prm}
    for gi in g:
        try:
            (_, s_cur), (_, p_cur) = gi.items()
            assert _ == 'p'  # parameters are not interchanged
            for p_id in prm:
                if s_cur == p_id:
                    if p_detected[p_id]:
                        lf.warning('Have multiple messages for {}!', prm[p_id])  # never occurred
                        p[p_id] += p_cur
                    else:
                        p_detected[p_id] = True
                        p[p_id] = p_cur
        except:
            print(gi), print(g)
            continue

    if len(p[id_coords]) == 0:  # no data
        return ()

    if not all(p_detected.values()):
        for p_id, b_detected in p_detected.items():
            if not b_detected:
                lf.warning('Not loaded: {}!!!', prm[p_id]['cols'])
                if p_id == id_coords:  # main data absent
                    raise Ex_nothing_done

    # Show device settings
    try:
        config_last = p.pop(id_config)[-1]
        lf.info('Last config:\n{}', config_last['p'][0].replace(' |', '\n'))
    except IndexError:
        pass  # Not fatal to not display info

    del prm[id_config]  # not for output to DataFrame, no more needed
    dfs = []
    for p_id, p_cfg in prm.items():  # All parameters to ou
        p_cols = p_cfg['cols']
        # 'r' - device time, 's' - receive time, 'p' - data
        df_str = pd.DataFrame.from_records(p[p_id], index='r', exclude='s')['p']
        p_list = df_str.tolist()
        # old: drop without pandas: if 'drop' in p_cfg:
        # p_cols_keep = [c for c in p_cols if c not in p_cfg['drop']]
        # i_keep = [i for i, c in enumerate(p_cols) if c not in p_cfg['drop']]
        # p_list = [[pi[i] for i in i_keep] for pi in p_list]

        df = pd.DataFrame.from_records(
            p_list,  # to_numpy(dtype=), #.
            index=df_str.index,
            columns=p_cols,
            exclude=p_cfg.get('drop')
            )

        if 'comma_to_dot' in p_cfg:
            for c in p_cfg['comma_to_dot']:
                if p_cfg['types'] == np.int8:
                    # comma and all after is not needed:
                    df[c] = df[c].str.replace(',.*', '', regex=True)
                else:
                    df[c] = df[c].str.replace(',', '.')

        df = df.astype(p_cfg['types'], copy=False)

        df_len = len(df)
        if p_id == id_coords:  # 1st params group
            df1st = df
        else:
            dfs.append(df)
            dlen = df_len_prev - df_len
            # on merge data without pair will be ignored. Warning if more than 1 points:
            (lf.warning if dlen > 1 else lf.debug)(
                'Data lengths different, will delete {} row of ({}) > ({})',
                dlen,
                *(lambda x, y: (x, y) if dlen > 0 else (y, x))(
                    ','.join(dfs[-1].columns.tolist()), ','.join(p_cols)
                    )
                )
        df_len_prev = df_len

    # as I see indexes values of dataframes are equal so merge_asof() is not needed
    nav_df = df1st.join(dfs, how='inner')  # preferred 'left' will change int dtypes to float if dfs lengths different!
    nav_df.index = pd.to_datetime(nav_df.index, unit='s', utc=True, origin=dt_from_utc.total_seconds())
    nav_df.index.name = 'Time'

    nav_df['LGSM'] = -nav_df['LGSM']
    # if nav_df.dtypes['LGSM']
    #     nav_df['LGSM'].astype(np.uint8)
    return nav_df


def loading(
        table: str,
        path_raw_local: Union[str, Path],
        time_interval: List[pd.Timestamp],
        dt_from_utc: timedelta) -> pd.DataFrame:
    """
    Loads Autofon/Spot data from xlsx or Autofon server.
    Function works only for known devices listed in globdl tables2mid
    input config:
    :param table: str
    :param path_raw_local: if no then loads from Autofon server and converts to DataFrame by autofon_df_from_dict()
    :param time_interval: 2 elements list of tz-aware pandas Timestamp
    :param dt_from_utc:
    :return: pandas DataFrame
    """
    mid = tables2mid[table]
    device_type, device_number = re.match(r'.*(sp|tr)#?(\d*).*', table).groups()
    if path_raw_local:
        if device_type =='sp':  # satellite based tracker
            if isinstance(path_raw_local, str):
                path_raw_local = Path(path_raw_local)
            if path_raw_local.suffix == '.xlsx':
                xls = pd.read_excel(path_raw_local,
                                    sheet_name=f'{device_number} Positions & Events',
                                    usecols='B,D',
                                    skiprows=4,
                                    index_col=0)
                nav_df = xls['Lat/Lng'].str.extract(r'(?P<Lat>[^,]*), (?P<Lon>[^,]*)')
                nav_df.set_index((nav_df.index - dt_from_utc).tz_localize('utc'), inplace=True)
                for coord in ['Lat', 'Lon']:
                    nav_df[coord] = pd.to_numeric(nav_df[coord], downcast='float', errors='coerce')
            else:
                # Load list_of_lists from mailbox
                time_lat_lon = spot_tracker_data_from_mbox(path_raw_local,
                                subject_end=f': {device_number}',
                                time_start=time_interval[0])
                nav_df = pd.DataFrame(time_lat_lon,
                                      columns=['Time', 'Lat', 'Lon']).astype(
                                        {'Lat': np.float32, 'Lon': np.float32}
                                        )
                nav_df['Time'] -= dt_from_utc
                nav_df.set_index('Time', inplace=True)  # pd.DatetimeIndex(nav_df['Time']
                nav_df.sort_index(inplace=True)
                nav_df = nav_df.tz_localize('utc', copy=False)

        else:
            prepare_loading_xlsx_links_by_pandas()
            xls = pd.read_excel(path_raw_local, usecols='C:D', skiprows=4, index_col='Дата')
            nav_df = xls['Адрес / Координаты'].str.extract(r'[^=]+\=(?P<Lat>[^,]*),(?P<Lon>[^,]*)').astype(
                                        {'Lat': np.float32, 'Lon': np.float32}
                                        )
            nav_df.set_index((nav_df.index - dt_from_utc).tz_localize('utc'), inplace=True)
            # for coord in ['Lat', 'Lon']:
            #     nav_df[coord] = pd.to_numeric(nav_df[coord], downcast='float', errors='coerce')


        nav_df = nav_df.truncate(*time_interval, copy=False)  # [k] - dt_from_utc for k in [0,1] .tz_localize('utc')
        tim_last_coord = pd.Timestamp.now(tz='utc')
    elif device_type == 'sp':  # satellite based tracker
        # download and parse from gmail
        time_lat_lon = spot_from_gmail(
            device_number=device_number,
            time_start=time_interval[0] - dt_from_utc)
        nav_df = pd.DataFrame(time_lat_lon,
                              columns=['Time', 'Lat', 'Lon']).astype(
            {'Lat': np.float32, 'Lon': np.float32}
            )
        nav_df['Time'] -= dt_from_utc
        nav_df.set_index('Time', inplace=True)  # pd.DatetimeIndex(nav_df['Time']
        nav_df.sort_index(inplace=True)
        nav_df = nav_df.tz_localize('utc', copy=False)
    else:
        url = 'http://176.9.114.139:9002/jsonapi'
        key_pwd = 'key=d7f1c7a5e53f48a5b1cb0cf2247d93b6&pwd=ao.korzh@yandex.ru'

        # Request and display last info
        r = requests.post(f'{url}/laststates/?{key_pwd}')
        if r.status_code != 200:
            print(r)
            raise(Ex_nothing_done)
        for d in r.json():
            if d['id'] != mid:
                continue
            tim_last_coord = pd.Timestamp(datetime.fromtimestamp(d['tscrd']), tz=timezone(dt_from_utc))
            lf.info(f'Last coordinates on server for #{{id}}: {tim_last_coord}, (N{{lat}}, E{{lng}})'.format_map(d))

        # Request and save new data

        if False:  # this will obtain filtered data useful for display only
            r = requests.post(f'{url}/?{key_pwd}',
                              json=[{'mid': str(k), **time_interval} for k in tables2mid.keys()]
                              )
            if r.status_code != 200:
                print(r)
                raise(Ex_nothing_done)
            nav_df = pd.DataFrame.from_records(d['points'], index='ts') \
                .rename(columns={'lat': 'Lat', 'lng': 'Lon', 'v': 'Speed'})

        r = requests.post(
            f"{url}/messages/{mid}?{key_pwd}&fromdate={{}}&todate={{}}&minlevel=6".format(
                *[int(t.timestamp()) for t in time_interval]))  # &lang=en not works
        if r.status_code != 200:
            print(r)
            raise(Ex_nothing_done)
        nav_df = autofon_df_from_dict(r.json()[0]['items'], dt_from_utc)

    if len(nav_df) == 0:
        lf.info(f"Downloaded {len(nav_df)} points")
        # have all data already => return without error: if no new data and time from last download is less than minute
        if time_interval[1] - tim_last_coord.tz_convert('utc') > timedelta(minutes=1):
            lf.warning(f"No data interval: {time_interval[1] - tim_last_coord.tz_convert('utc')}")
        raise(Ex_nothing_done)

    # display last point with local time
    lf.info(f"Downloaded {len(nav_df)} points for #{mid}, last - "
            f"{nav_df.index[-1].tz_convert(timezone(dt_from_utc))}: "
            "{Lat:0.6f}N, {Lon:0.6f}E".format_map(nav_df.iloc[-1]))
    return nav_df


def saving(nav_dfs, path, process):
    for tbl, nav_df in nav_dfs.items():
        save2gpx(nav_df, tbl, path=path, process=process)

        # bin average data
        bins = ['1H', '5min']
        for bin in bins:
            save2gpx(nav_df.resample(bin).mean(), f'{tbl}_avg({bin})', path=path, process=process)


def proc(cfg=cfg):
    cur_time = datetime.now()
    time_interval = {
        'start': int((cur_time - timedelta(days=100)).timestamp()),
        # 1st data to load, will be corrected if already have some
        'end': int(cur_time.timestamp())
        }

    nav_dfs = []
    for tbl in cfg['in']['tables']:
        nav_dfs[tbl] = call_with_valid_kwargs(loading,
                                              time_interval=time_interval,
                                              **cfg['in']
                                              )
    saving(nav_dfs,
           path=cfg['out']['path'],
           process=cfg['process']
           )


class OpenHDF5(FakeContextIfOpen):
    """
    Context manager that do nothing if file is not str/PurePath or custom open function is None/False
    useful if instead file want use already opened file object
    Returns table names and opened store

    :param tables: tables names search pattern or sequence of table names
    :param tables_log: tables names for metadata of data in `tables`
    :param db_path:
    :return: iterator that returns (table name, coefficients)
    updates cfg_in['tables'] - sets to list of found tables in store
    """
    # will be filled by each table from cfg['in']['tables']

    def __init__(self, db_path, tables, tables_log, db=None):
        self.tables = tables
        self.tables_log = tables_log
        self.handle = super(OpenHDF5, self).__init__(
            lambda f: pd.HDFStore(f, mode='r'),
            file=db_path,
            opened_file_object=db)

    def __enter__(self):
        """
        :return: opened handle or :param file: from __init__ if not need open
        """
        self.handle = super(OpenHDF5, self).__enter__()
        return (
            self.handle,
            self.tables,
            self.tables_log or [f'{t}/log' for t in self.tables]
            )

hydra.output_subdir = 'cfg'
# hydra.conf.HydraConf.output_subdir = 'cfg'
# hydra.conf.HydraConf.run.dir = './outputs/${now:%Y-%m-%d}_${now:%H-%M-%S}'


@dataclass
class ConfigInAutofon:
    time_interval: List[str] = field(default_factory=lambda: ['2021-04-08T12:00:00', 'now'])
    # use already loaded coordinates instead of request:
    path_raw_local: Optional[str] = None
    dt_from_utc_hours: int = 0
    # b_incremental_update: bool = True


@dataclass
class ConfigProcessAutofon:
    # gpx track settings
    simplify_tracks_error_m = 0
    dt_per_file_days = 356
    b_missed_coord_to_zeros: bool = False
    period_tracks: Optional[str] = None
    period_segments: Optional[str] = '1D'
    # anchor settings
    anchor_coord: List[float] = field(default_factory=lambda: [44.56905, 37.97308])
    anchor_coord_time_dict: Dict[Any, Any] = field(default_factory=dict)  # {time: [Lat, Lon]} - use if anchor moved
    anchor_depth: float = 0
    max_dr: float = 100  # Max distance to anchor, m. Delete data with dr > max_dr

    # absent data settings
    # detect absent data to try download again. Default is '10min' for gprs and '20min' for satellite based tracker:
    dt_max_hole: Optional[str] = None
    # detect absent data only in this interval back from last data. Default is '1D' for gprs and '0D' for satellite based tracker:
    dt_max_wait: Optional[str] = None

    # other
    b_reprocess: bool = False  # reprocess all data (previously calculated dx, dy...) from Lat, Lon, DateTimeIndex



@dataclass
class ConfigOutAutofon(cfg_dataclasses.ConfigOutSimple):
    # dt_bins_rolling: List[List[str]] = field(default_factory=lambda: [['2H', None], ['5min', None], ['10min', '1H']])
    # List[List[ or List[Optional[str] not supported so we split it:
    dt_bins: List[str] = field(default_factory=lambda: ['2H', '5min', '10min'])
    dt_rollings: List[str] = field(default_factory=lambda: ['', '', '1H'])
    # len(to_gpx) should be less than 1(for raw table) + dt_bins, if empty then output tho gpx if averaging less than 1h.
    to_gpx: List[bool] = field(default_factory=list)

ConfigProgram = cfg_dataclasses.ConfigProgram

cs_store_name = Path(__file__).stem
cs, ConfigType = cfg_dataclasses.hydra_cfg_store(
    cs_store_name, {
    'input': ['in_autofon'],  # Load the config "in_hdf5" from the config group "input"
    'out': ['out_autofon'],  # Set as MISSING to require the user to specify a value on the command line.
    #'filter': ['filter'],
    'process': ['process_autofon'],
    'program': ['program'],
    # 'search_path': 'empty.yml' not works
    },
    module=sys.modules[__name__]
    )


def dx_dy_dist_bearing(lon1, lat1, lon2, lat2):
    lon1, lat1, lon2, lat2 = map(np.radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1

    R = 6371000  # radius of the earth in m
    klon = np.cos((lat2+lat1)/2)
    dx = R * klon*dlon
    dy = R * dlat

    d = np.sqrt(dy**2 + dx**2)
    angle = np.arctan2(dlat, dlon)
    return np.column_stack((dx, dy, d, np.degrees(angle)))

    #haversine_
    # x = sin(dlon) * cos(lat2)
    # y = cos(lat1) * sin(lat2) - (sin(lat1) * cos(lat2) * cos(diffLong))
    #
    #
    # m = 6367000 * 2 * np.arcsin(np.sqrt(a))
    #
    # dx = np.cos(lat1) * np.cos(lat2) * np.sin(dlon/2)
    # dy =np.sin(dlon / 2.0)
    # a = np.sin(dlat / 2.0)


def format_log_filename(start, end):
    return '{:%y%m%d_%H%M}-{:%m%d_%H%M}'.format(start, end)


def proc_and_h5save(df, tbl, cfg_in, out, process, bin: Optional[str] = None, rolling_dt: Optional[str] = None):
    """
    Calculates displacement, bin average and saves to HDF5 and GPX
    For averaged data tables if Course column exist then from Course and dr calculates averaged 'speed_x' and 'speed_y'
     with dropping 'Course'.
    :param df: DataFrame with datetime index and columns:
        Lat, Lon: coordinates
    :param tbl: table name where to save result in Pandas HDF5 store
    Configuration dicts:
    :param cfg_in:
    :param out: dict, output config
    :param process: dict, processing config
    Averaging parameters:
    :param bin:
    :param rolling_dt:
    :return:
    Modifies:
     df: columns of polar data and what will be calculated after average are removed,
     out['table']=tbl,
     out['tables_written'] in h5_sort_filt_append()
    """
    if not any(df):
        return ()

    out['log']['fileChangeTime'] = pd.Timestamp.now()  # :%y%m%d_%H%M%S
    out['log']['fileName'] = format_log_filename(*df.index[[0, -1]])
    del out['log']['index']  # was temporary used internally

    if bin is not None:
        # Drop Vdir and Course because it is not correct to average angles: will be calculated after average along with
        # other columns: ['dx', 'dy', 'dr', 'Vdir'] from <Lat>, <Lon>.
        # 'Course' we need to be convert to dekart: For vector length we using `dr` instead `Speed` which is too rough
        # dekart2polar_df_uv(df.rename())

        if 'Course' in df.columns:
            # if already done then no 'dx' because of side effects here, same as 'Course' (to not repeat or try if no Course)
            # reuse 2 float columns instead of dropping, then drop other:
            df.rename(columns={'dx': 'speed_x', 'dy': 'speed_y'}, inplace=True)
            course = np.radians(df['Course'].values)

            df['speed_x'], df['speed_y'] = df['dr'].values * np.vstack([np.sin(course), np.cos(course)])
            df.drop(['dr', 'Vdir', 'Course'], axis='columns', inplace=True)

        # bin average data
        shift_to_mid = to_offset(bin) / 2
        df = df.resample(bin, offset=-shift_to_mid).mean()
        df.index += shift_to_mid

        if rolling_dt:
            window_not_even = int(re.match('.*mov(\d+)bin', tbl).group(1))
            lf.info('{}: moving average over {} bins of {}', tbl, window_not_even, bin)
            out['log']['fileName'] += f'bin{bin}'
            df = df.resample(bin, origin='epoch').mean().rolling(
                window_not_even, center=True, win_type='gaussian', min_periods=1).mean(std=3)
            out['log']['fileName'] += f'avg{rolling_dt}'
        else:
            lf.info('{}: {}-bin average', tbl, bin)
            out['log']['fileName'] += f'avg{bin}'

    df.dropna(subset=['Lat', 'Lon'], inplace=True)  # *.gpx will be not compatible to GPX if it will have NaN values

    if process['anchor_coord_time_dict']:
        anchor_lat_lon = np.zeros((2, len(df)), np.float32)
        if any(process['anchor_coord']):
            anchor_lat_lon += np.float32(process['anchor_coord'])[:, None]
        anchor_times = pd.DatetimeIndex(process['anchor_coord_time_dict'].keys(), tz='utc')
        i_starts = np.searchsorted(df.index, anchor_times).tolist()
        for i_st, i_en, lat_lon in zip(i_starts, i_starts[1:] + [len(df)], process['anchor_coord_time_dict'].values()):
            anchor_lat_lon[:, i_st:i_en] = np.float32(lat_lon)[:, None]
        anchor_coord = anchor_lat_lon[::-1, :]
    else:
        anchor_coord = process['anchor_coord'][::-1] if any(process['anchor_coord']) else 0

    # Calculate parameters
    df.loc[:, ['dx', 'dy', 'dr', 'Vdir']] = dx_dy_dist_bearing(
        *anchor_coord,
        *df[['Lon', 'Lat']].values.T
        )
    # course, azimuth2, distance = geod.inv(  # compute forward and back azimuths, plus distance
    #     *df.loc[navp_d['indexs'][[0, -1]], ['Lon', 'Lat']].values.flat)  # lon0, lat0, lon1, lat1

    # Filter source data
    if bin is None and process['max_dr']:
        b_bad = df.dr < process['max_dr']
        b_bad_sum = b_bad.sum()
        if b_bad_sum:
            lf.warning('{} rows with dr < {} deleted', b_bad_sum, process['max_dr'])
            df = df[b_bad]

    # Saving to HDF5
    out['table'] = tbl
    h5_sort_filt_append(df, input={**cfg_in, 'dt_from_utc': timedelta(0)}, out=out)
    return df


def holes_starts(t: pd.DatetimeIndex, t_max: int) -> Tuple[pd.DatetimeIndex, np.timedelta64]:
    """
    Finds time starts of data holes and its sizes
    :param t: time data
    :param t_max: max time difference considered not to be hole
    :return: t_start, dt
    """
    dt_int = np.ediff1d(t.view(np.int64), to_end=0)  # or .to_numpy(dtype=np.int64)
    i_hole = np.flatnonzero(dt_int > t_max)
    dt = np.array(dt_int[i_hole], 'm8[ns]').astype('m8[s]')  # np.round(dt_int[i_hole]*1E-9, 1)
    t_start = t[i_hole]
    n_holes = len(t_start)
    if n_holes:
        n_show = 20
        msg_number = f'Last {n_show} of {n_holes}' if n_holes > n_show else f'Found {n_holes}'
        lf.warning('{} holes:\n{}', msg_number,
            tabulate({'time_start': t_start[-n_show:],
                      'dt_minutes': dt.astype(int)[-n_show:] / 60},
                     headers='keys',
                     floatfmt='.1f'
                     )
                   )
    else:
        lf.debug('no holes found')
    return t_start, dt


def holes_prepare_to_fill(db, tbl, tbl_log,
                          time_holes: Optional[List[pd.Timestamp]] = None,
                          dt_max_hole: Optional[str] = '10min',
                          dt_max_wait: Optional[str] = '1D'
                          ) -> Tuple[Optional[List[pd.Timestamp]], Optional[pd.Timestamp], Optional[str]]:
    """
    Finds time holes in data index to download on data's hole start (if time_holes is None else uses it)

    :param db: opened pandas HDF5 data store
    :param tbl: table name
    :param tbl_log: log table name
    Search holes parameters:
    :param time_holes:  None to search holes. If DatetimeIndex then return it as time data gaps.  Will select 1st gap > max_time - dt_max_wait
    :param dt_max_hole: Search holes of this size at least
    :param dt_max_wait: Search/select holes only in this range from end:
        None: search everywhere/not change
        '0': do not use but search for 1 day and display message
    :return:
     time_holes: list of found holes' starts
     time_start: time of last data if no holes we want to fill else hole start (starting Timestamp to query server)
     msg: log info of reason for time_start used: 'last hole' or 'last saved'
    """

    try:  # Last data time from log table
            t_max_exist = db.select(tbl_log, columns=['DateEnd'], start=-1)['DateEnd'][0]
    except (KeyError, IndexError):      # no log (yet or lost?)
        try:
            t_min_exist = db.select(tbl, columns=[], stop=1).index[0]
            t_max_exist = db.select(tbl, columns=[], start=-1).index[0]
            lf.warning('Log lost! Appending one row for all data: {} - {}', t_min_exist, t_max_exist)
            # try:
            #     t_date_end = t_max_exist.tz_convert(db.select(tbl_log, columns=['DateEnd']).dtypes['DateEnd'].tz)
            # except:
            #     lf.exception('old format converting fail but may be not needed if it is new')
            #     t_date_end = t_max_exist  # to try just as is

            df_log = pd.DataFrame(
                {'DateEnd': [t_max_exist],
                 'fileChangeTime': np.datetime64('now'),
                 'fileName': format_log_filename(t_min_exist, t_max_exist),
                 'rows': -1},
                index=[t_min_exist]
            )
            h5.append_log(df_log, tbl_log, {'db': db, 'nfiles': None})
        except (KeyError, IndexError, AttributeError):  # no data yet  # not need if all worked properly
            return time_holes, None, 'start'


    time_start_wait = t_max_exist - to_offset(dt_max_wait).delta if dt_max_wait else None
    max_hole_timedelta = to_offset(dt_max_hole)

    if time_holes is None:
        # searching holes
        for try_query in (
                        [f"index > ('{time_start_wait}')", ''] if time_start_wait else
                        [''] if time_start_wait is None else
                        ['1D']):
            try:
                time_holes, dt_holes = holes_starts(db.select(tbl, try_query, columns=[]).index,
                                                    max_hole_timedelta.nanos)
                break
            except Exception as e:  # tables.exceptions.NoSuchNodeError  # have only once, may be not needed
                if try_query:       # 'UnImplemented' object has no attribute 'description'
                    lf.error('Can not query {} for "{}": {}. Retrying in memory', tbl, try_query, e)
                    continue  # Checking for holes all data
                else:
                    raise HDF5ExtError('Can not load data from "{}/{}"'.format(db, tbl))
    elif time_start_wait:
        time_holes = time_holes[time_holes >= time_start_wait]

    if len(time_holes) and time_start_wait:
        msg_start_origin = 'last hole'
        t_start = time_holes[0]

        # try:
        #     df_log_cur = db.select(
        #         tbl_log,
        #         f"index <= Timestamp('{t_start}') & "
        #         f"DateEnd > Timestamp('{t_start}')")
        # except Exception as e:  # tables.exceptions.NoSuchNodeError  # have only once, may be not needed
        #     lf.error('Can not query {} for "{}": {}. Doing in memory', tbl_log, try_query, e)
        #     df_log_cur = db[tbl_log]
        #     b_good = (df_log_cur.index <= t_start) & (t_start < df_log_cur.DateEnd)
        #     df_log_cur = df_log_cur[b_good]
        #
        # # log_dict['fileName'] =
        # # log_dict['fileChangeTime'] = df_log_cur['fileChangeTime'].iat[0]
        #
        # if len(df_log_cur) != 1:                # duplicates, how?
        #     df_log_cur = df_log_cur.iloc[[0]]   # to del. next rows = duplicates
        # if not h5.rem_last_rows(db, [tbl, tbl_log], [df_log_cur], t_start):
        #     return time_holes, None, 'new start'            # failed to use previous data
    else:
        msg_start_origin = 'last saved'
        t_start = t_max_exist
    return time_holes, t_start, msg_start_origin

    # df_log[df_log['fileName'] == log['fileName']]
    # df_log = cfg_out['db'][tbl_log]
    # t_max_exist = df_log['DateEnd'].max()
    #
    #
    # qstr = "index>=Timestamp('{}')".format(df_log_cur.index[0])
    # qstrL = "fileName=='{}'".format(df_log_cur['fileName'][0])
    #
    # # df = cfg_out['db'][tbl]
    # df = cfg_out['db'].select(tbl, where=qstr)


def h5.names_gen(cfg_in: Mapping[str, Any],
                 cfg_out: MutableMapping[str, Any],
                 processed_tables: bool,
                 dt_max_hole: Optional[str] = None,
                 dt_max_wait: Optional[str] = None,
                 **kwargs) -> Iterator[Union[
        Tuple[str, str, pd.Timestamp, str, bool],
        Tuple[str, str, str, str, bool]
        ]]:
    """
    Generate table names from cfg_out['tables'], assigns required log fields to cfg_in['time_interval']
    1st is raw table and other are processed averaged as defined by 'dt_bins' and 'dt_rollings'
    :param cfg_in:
    :param cfg_out:
    :param processed_tables: function will work in mode of generating parameters to average already loaded data
    :param dt_max_hole: holes_prepare_to_fill() parameter:
    :param dt_max_wait: holes_prepare_to_fill() parameter:
    :param kwargs:
    # time_holes: holes_prepare_to_fill() parameter: list of found holes' starts
    # time_start: holes_prepare_to_fill() parameter: starting Timestamp to query server
    :return: (tbl, tbl_log, t_start, msg_start_origin, to_gpx) or
             (tbl, tbl_log, bin, rolling_dt, to_gpx):
    (`t_start`, msg_start_origin) replaced with (`bin`, rolling_dt) if processed_tables
    t_start, msg_start_origin: see holes_prepare_to_fill()
    """
    set_field_if_no(cfg_out, 'log', {})
    # tables_log_copy = cfg_out['tables_log']
    for tbl in cfg_out['tables']:
        cfg_out['log']['fileName'], cfg_out['log']['fileChangeTime'] = cfg_in['time_interval']
        tbl_log = cfg_out['tables_log'][0].format(tbl)
        # cfg_out['tables_log'] = [tbl_log]
        try:
            if not processed_tables:
                # mode 1: check existed data for holes and output its last good time: t_start
                for retry in [False, True]:
                    try:
                        time_holes, t_start, msg_start_origin = holes_prepare_to_fill(cfg_out['db'], tbl, tbl_log,
                                                                          time_holes=kwargs.get('time_holes', None),
                                                                          dt_max_hole=dt_max_hole,
                                                                          dt_max_wait=dt_max_wait)
                    except HDF5ExtError:
                        lf.exception('Bad DB table. Recovering table...')
                        try:
                            h5.remove_tables(cfg_out['db'], [tbl], [])
                            cfg_out['db'].close()
                        except:
                            lf.exception('Bad DB. Recovering table failed. Recovering DB.'
                                         'Deleting temporary to replacing by output store...')
                            cfg_out['db'].close()
                            cfg_out['temp_db_path'].unlink()
                        try:
                            h5.move_tables(cfg_out, tbl_names=[tbl])
                        except KeyError:
                            lf.exception('no data?')
                            t_start = None
                            msg_start_origin = 'beginning again'
                            retry = False
                        cfg_out['db'] = pd.HDFStore(cfg_out['temp_db_path'])
                        if retry:
                            lf.exception('Bad DB. Deleting failed')
                            raise
                        else:
                            continue

                cfg_out['log']['index'] = t_start  # will be read in h5.del_obsolete()

                yield (tbl, tbl_log, t_start, msg_start_origin, cfg_out['to_gpx'][0])
                # cfg_out['tables_log'] = tables_log_copy
            else:
                # mode 2: Generate parameters for tables that will be processed
                # todo: load last data time - averaging interval shift, check last data for each table
                t_start = cfg_in.get('t_good_last')
                for bin, rolling_dt, to_gpx in zip_longest(
                        cfg_out['dt_bins'],
                        cfg_out['dt_rollings'],
                        cfg_out['to_gpx'][1:]
                        ):
                    # Table name form averaging parameters
                    bin_timedelta = to_offset(bin)
                    if rolling_dt:
                        rolling_timedelta = to_offset(rolling_dt)
                        window_not_even = int(rolling_timedelta / bin_timedelta)
                        if not (window_not_even % 2):
                            window_not_even += 1
                        tbl_avg = f'{tbl}_avg{rolling_dt.lower()}=mov{window_not_even}bin{bin.lower()}'
                        max_timedelta = max(bin_timedelta, rolling_timedelta)
                    else:
                        tbl_avg = f'{tbl}_avg{bin.lower()}'
                        max_timedelta = bin_timedelta

                    # default not output to gpx if averaging is bigger than 1 hours
                    if to_gpx is None:
                        to_gpx = max_timedelta <= timedelta(hours=1)
                    cfg_out['log']['index'] = t_start  # will be read in h5.del_obsolete()
                    tbl_log = cfg_out['tables_log'][0].format(tbl_avg)
                    cfg_tables_save, cfg_out['tables'] = cfg_out['tables'], [tbl_avg]  # for compatibility with del_obsolete()
                    yield (tbl_avg, tbl_log, bin, rolling_dt, to_gpx)
                    cfg_out[tables] = cfg_tables_save                              # recover
        except GeneratorExit:
            print('Something wrong?')
            return ExitStatus.failure


def h5move_and_sort(out: MutableMapping[str, Any]):
    """
    Moves from temporary storage and sorts `tables_written` tables and clears this list
    :param out: fields:
        tables_written: set of tuples of str, table names
        b_del_temp_db and other fields from h5.index_sort
    :return:
    Modifies out fields:
        b_remove_duplicates: True
        tables_written: assigns to empty set

    """

    failed_storages = h5.move_tables(out, tbl_names=out['tables_written'])
    print('Finishing...' if failed_storages else 'Ok.', end=' ')
    # Sort if have any processed data, else don't because ``ptprepack`` not closes hdf5 source if it not finds data
    out['b_remove_duplicates'] = True
    h5.index_sort(out,
                 out_storage_name=f"{out['db_path'].stem}-resorted.h5",
                 in_storages=failed_storages,
                 tables=out['tables_written']
                 )
    out['tables_written'] = set()


@hydra.main(config_name=cs_store_name, config_path="cfg", version_base='1.3')  # adds config store cs_store_name data/structure to :param config data/structure to :param config
def main(config: ConfigType) -> None:
    """
    ----------------------------
    Save data tp GPX files and Pandas HDF5 store*.h5
    ----------------------------
    The store contains tables for each device and each device table contains log with metadata of recording sessions

    :param config: with fields:
    - in: mapping with fields:
      - tables_log: - log table name or pattern str for it: in pattern '{}' will be replaced by data table name

    - out: mapping with fields:

    """
    if GetMutex().IsRunning():
        lf.info("Application is already running. Waiting 10s...")
        sleep(10)
        sys.exit(ExitStatus.failure)

    global cfg
    cfg = cfg_dataclasses.main_init(config, cs_store_name)
    cfg_in = cfg.pop('input')
    cfg_in['cfgFile'] = cs_store_name
    cfg['in'] = cfg_in
    # try:
    #     cfg = to_vaex_hdf5.cfg_dataclasses.main_init_input_file(cfg, cs_store_name, )
    # except Ex_nothing_done:
    #     pass  # existed db is not mandatory
    # geod = pyproj.Geod(ellps='WGS84')
    # dir_create_if_need(out['text_path'])
    out = cfg['out']
    # if relative path for cfg['out']['db_path'] then it is from directory of running script
    if cfg['out'].get('db_path') and not cfg['out']['db_path'].is_absolute():
        cfg['out']['db_path'] = Path(sys.argv[0]).parent / cfg['out']['db_path']

    h5.out_init(cfg['in'], out)

    if not out['to_gpx']:  # default to output to gpx raw data and for averaged only if averaging is less than 1 hours
        out['to_gpx'] = [True]  # output 1st, other (will be set to None) depends on averaging

    load_from_internet = True   # for debug: allows to not load
    out['tables_written'] = set()

    # device specific default parameters to check existed data
    b_sp = all('sp' in t for t in out['tables'])
    if cfg['process']['dt_max_hole'] is None:
        cfg['process']['dt_max_hole'] = '20min' if b_sp else '10min'
    if cfg['process']['dt_max_wait'] is None:
        cfg['process']['dt_max_wait'] = '0D' if b_sp else '1D'

    out['b_incremental_update'] = True              # enables to determine last data we have
    out['field_to_del_older_records'] = 'index'     # new data priority (based on time only)
    out['b_reuse_temporary_tables'] = True          # reuse previous temporary data
    msg_start_fmt = '{} {} {:%y-%m-%d %H:%M:%S%Z} \u2013 {:%m-%d %H:%M:%S%Z}{}'
    time_interval_default = [pd.Timestamp(t, tz='utc') for t in cfg_in['time_interval']]

    for b_retry in [False, True]:
        # Loading data for each probe to temporary store cycle
        ######################################################
        for i1_tbl, (tbl, tbl_log, t_start, msg_start_origin, _) in h5.dispenser_and_names_gen(
                cfg_in, out,
                fun_gen=h5.names_gen,
                processed_tables=False,
                dt_max_hole=cfg['process']['dt_max_hole'],
                dt_max_wait=cfg['process']['dt_max_wait']
                ):
            lf.info('{}. {}: ', i1_tbl, tbl)

            time_interval = time_interval_default
            if load_from_internet:
                # Time interval for new data loading query
                if t_start is None:  # previous data last time not found - Ok, may be we have no data yet
                    lf.info(msg_start_fmt, 'Downloading', tbl, *time_interval, '...')
                else:
                    time_interval[0] = t_start
                    if msg_start_origin == 'last saved':
                        lf.info(msg_start_fmt, 'Continue downloading', tbl, *time_interval, f' (from {msg_start_origin})')
                        time_interval[0] += timedelta(seconds=1)  # else 1st downloaded row will be same as the last we have
                    elif t_start:  #  'last hole'
                        lf.warning(msg_start_fmt, 'Continue downloading', tbl, *time_interval,
                                   f': filling gaps. {time_interval_default[-1] - t_start} from {msg_start_origin} to now!')
                    else:  # possible?
                        lf.error('t_start = {}!', t_start)

                # Loading data from internet
                try:
                    df_loaded = call_with_valid_kwargs(loading, table=tbl, **{**cfg_in, 'time_interval': time_interval})
                except (Ex_nothing_done, TimeoutError):
                    sys.exit(ExitStatus.failure)

            # Reprocess option: df_loaded = previous source data + df_loaded. Then del previous data table.
            if cfg['process']['b_reprocess']:
                lf.info('prepend all previous stored source data in memory, and reprocess')
                df, df_log = load_prev_source_data(
                    tbl, tbl_log, out['db'], out['db_path'], select_where=f"index < ('{df_loaded.index[-1]}')")
                if df is not None:  # selected df.index < df_loaded.index[-1]
                    df_loaded = df[df_loaded.columns].append(df_loaded)
                    h5.remove_tables(out['db'], [tbl], [])
                    if df_log.empty:  # empty log is not normal but skip if so
                        lf.warning('log is empty')
                    else:
                        # with h5.ReplaceTableKeepingChilds([df], tbl, cfg, h5.append_log):
                        #     pass
                        h5.append_log(df_log, tbl_log, {**out, **{'nfiles': None}})
                        #h5.append({'table': tbl, **out}, [], df_log, log_dt_from_utc=cfg_in['dt_from_utc'])
            # Write to temporary store
            proc_and_h5save(df_loaded, tbl, cfg_in, out, cfg['process'])
        if out.get('db_is_bad') and not b_retry:  # recover
            lf.warning(f'Recovering temp db from *.copy.h5 because of "db_is_bad" flag have been set')
            try:
                # clean start
                copyfile(out['temp_db_path'].with_suffix('.copy.h5'), out['temp_db_path'])
                out['db_is_bad'] = False
                continue
            except:
                lf.exception(f'Recovering temp db from *.copy.h5 failed')
        sleep(2) # trying to reduce HDF5ExtError rate
        break

    # Save current data last good time

    # and tables for which it is valid to reduce searching good start on next run
    # None if cfg_changed else cfg_in.get('t_good_last')

    # Update output store
    #######################
    tables_written = out['tables_written']  # what to copy from temp store
    if (not load_from_internet) or any(h5.unzip_if_need(tables_written)):  # cfg_in.get('time_last'):

        try:
            if out.get('db_is_bad'):
                db_path_temp_copy = h5.replace_bad_db(out['temp_db_path'], out['db_path'])
                lf.warning('Trying copy {} from {} file to {}', out['tables'], db_path_temp_copy, out['temp_db_path'])
                h5move_and_sort(
                    {'db_path': out['temp_db_path'],
                     'temp_db_path': db_path_temp_copy,
                     'tables_written': out['tables']
                     })
                sleep(2)
            h5move_and_sort(out)
        except HDF5ExtError:
            if out.get('db_is_bad'):
                lf.exception('Temp store is bad. Nothing helps to delete. Delete it manually!')
            else:
                h5.replace_bad_db(out['temp_db_path'], out['db_path'])
            sys.exit(ExitStatus.failure)
            # copy what we can: rename file then copy objects back

            # for retry in [False, True]:
            #     try:
            #         out['temp_db_path'].rename(db_path_temp_copy)
            #     except PermissionError:
            #         # if not out.get('db_is_bad'):
            #         if out['db'] and out['db'].is_open:
            #             out['db'].close()
            #         else:
            #             print('Wait store closing...')
            #             out['db'] = None
            #             gc.collect()
            #             sleep(1)
            #             continue
            #     break




            # with pd.HDFStore(out['temp_db_path'], mode='r') as store:
            #     df, df_log = load_prev_source_data(
            #         tbl, tbl_log, out['db'], store, select_where=f"index < Timestamp('{df_loaded.index[-1]}')")
            #     # . Replacing temporary store with copy of old result store
            # copyfile(out['db_path'], out['temp_db_path'])


            # if
            # need replace

        # Loading concatenated current and previous data. todo: only needed time range
        b_all_needed_data_in_temp_store = False
        with pd.HDFStore(out['db_path'], mode='r') as store_in:
            df_raw = store_in.select(tbl)  # we just saved to output (indexed) store so loading it from there

            # Overwriting temp store
            if not cfg_in.get('t_good_last'):
                # will process all data => not need older processed tables. Removing them by
                # replace temporary store with 1 result table (compressed, sorted and indexed) from result store:
                with tables.open_file(out['temp_db_path'], mode='w') as store_out:
                    store_in._handle.copy_node(
                        f'/{tbl}',
                        newparent=store_out.root,
                        recursive=True,
                        overwrite=True)
                    store_out.flush()  # .flush(fsync=True
                b_all_needed_data_in_temp_store = True
                # out_back = out
                # out_back['db_path'], out_back['temp_db_path'] = out_back['temp_db_path'], out_back['db_path']
                # out_back['db_path']
                # h5move_and_sort({'db_path': out[], 'tables_written': tables_written})

        if b_all_needed_data_in_temp_store:
            out['tables_written'] = tables_written
            out['db_path'].unlink()  # remove file

        print('\tSaving to GPX and other processing')
        if out['to_gpx'][0]:  # Saving GPX
            save2gpx(df_raw, tbl, path=out['db_path'], process=cfg['process'], dt_from_utc=cfg_in['dt_from_utc'])


        for i1_tbl, (tbl, tbl_log, bin, rolling_dt, b_to_gpx) in h5.dispenser_and_names_gen(
                cfg_in, out,
                fun_gen=h5.names_gen,
                processed_tables=True
                # dt_max_hole=cfg['process']['dt_max_hole'],
                # dt_max_wait=cfg['process']['dt_max_wait']
                ):
            # Calculate averages and save them to other tables in HDF5
            df = proc_and_h5save(df_raw, tbl, cfg_in, out, cfg['process'], bin=bin, rolling_dt=rolling_dt)
            # Saving GPX
            if b_to_gpx:
                save2gpx(df, tbl, path=out['db_path'], process=cfg['process'], dt_from_utc=cfg_in['dt_from_utc'])


        if any(h5.unzip_if_need(out['tables_written'])):
            # not need to remove temp db if want use out['b_reuse_temporary_tables'] option: but temporary tables remains else set out['b_del_temp_db'] = True
            h5move_and_sort(out)

    try:
        # replace temporary store with copy of (compressed, sorted and indexed) result store:
        if not out.get('db_is_bad'):
            copyfile(out['temp_db_path'], out['temp_db_path'].with_suffix('.copy.h5'))
        copyfile(out['db_path'], out['temp_db_path'])
    except:
        lf.exception('Can not replace temporary file {}', out['temp_db_path'])
    print(f"{datetime.now():%Y-%m-%d %H:%M:%S} Ok>", end=' ')


def load_prev_source_data(tbl: str, tbl_log: str, store: pd.HDFStore, db_path_other: Path, select_where=''):
    """
    Try load from df, df_log from tbl, tbl_log tables of store or from db_path_other
    :param tbl:
    :param tbl_log:
    :param select_where:
    :param store:
    :param db_path_other:
    :return: df, df_log

    """
    df = None
    df_log = None
    try:  # normally have data in opened temp store:
        df = store.select(tbl, where=select_where)
        df_log = store.select(tbl_log, where=select_where)
    except KeyError:  # 'No object named tr2 in the file': Was table occasionally deleted?
        lf.info('Nothing found to reprocess in temp db')
    if df is None or df_log is None or df_log.empty:
        try:  # if failed then load from output store:
            with pd.HDFStore(db_path_other, mode='r') as store_in:
                df = store_in.select(tbl, where=select_where)
                df_log = store_in.select(tbl_log, where=select_where)
        except OSError:  # ... .h5` does not exist
            lf.info('Nothing to reprocess in db. Not need load/remove')
    return df, df_log


def main_call(
        cmd_line_list: Optional[List[str]] = None,
        fun: Callable[[], Any] = main
        ) -> Dict:
    """
    Adds command line args, calls fun, then restores command line args
    :param cmd_line_list: command line args of hydra commands or config options selecting/overwriting

    :return: global cfg
    """

    sys_argv_save = sys.argv.copy()
    if cmd_line_list is not None:
        sys.argv += cmd_line_list

    # hydra.conf.HydraConf.run.dir = './outputs/${now:%Y-%m-%d}_${now:%H-%M-%S}'
    fun()
    sys.argv = sys_argv_save
    return cfg


if __name__ == '__main__':
    main()  # [f'--config-dir={Path(__file__).parent}'])


def pairwise(iterable):
    "s -> (s0, s1), (s2, s3), (s4, s5), ..."
    a = iter(iterable)
    return zip(a, a)


def call_example():
    """
    to run from IDE or from bat-file with parameters
    --- bat file ---
    call conda.bat activate py3.7x64h5togrid
    D: && cd C:/Work/Python/AB_SIO_RAS/h5toGrid
    python -c "from to_vaex_hdf5.autofon_coord import call_example; call_example()"
    ----------------
    # python -m to_vaex_hdf5.autofon_coord.call_example() not works
    :return:
    """
    # from to_vaex_hdf5.h5tocsv import main_call as h5tocsv
    path_db = Path(
        r'd:/workData/BalticSea/210515_tracker/current@tr0/210515_1500tr0.h5'
        # r'd:\WorkData\BlackSea\210408_trackers\tr0\210408trackers.h5'
        )
    device = ['tr0']  # 221912
    main_call([  # '='.join(k,v) for k,v in pairwise([   # ["2021-04-08T08:35:00", "2021-04-14T11:45:00"]'
        'input.time_interval=[2021-05-15T13:00:00, now]',  # ["2021-04-08T09:00:00", "now"]',   # UTC, max (will be loaded and updated what is absent)
        'input.dt_from_utc_hours=2',  #3
        'process.anchor_coord=[54.62425, 19.76050]', #[44.56905, 37.97309]',
        'process.anchor_depth=20',
        'process.period_tracks=1D',
        # 'process.period_segments="2H"', todo: not implemented to work if period_tracks is set
        f'out.db_path="{path_db}"',
        'out.tables=[{}]'.format(','.join([f'"{d}"' for d in device])),
        # 'out.tables_log=["{}/log"]',
        # 'out.b_insert_separator=False'
        # 'input.path_raw_local="{}"'.format({  # use already loaded coordinates instead of request:
        #     221910: Path('d:\Work') /
        #             r' координаты адреса выходов на связь 09-04-2021 08-03-13 09-04-2021 08-03-13.xlsx'
        #     }),
        # 'input.time_intervals={221910: "2021-04-08T12:00:00"}',

        ])

    # conda activate py3.7x64h5togrid && D: && cd C:\Work\Python\AB_SIO_RAS\h5toGrid && python -m to_vaex_hdf5.autofon_coord.call_example()

# r = requests.post(
#     'http://176.9.114.139/jsonapi/?key=5cee5887daa94dff88e51dc9bd6d16a2&pwd=0887219675',  # &a0ghzybxtr
#     json={'mid': 'ao', **time_interval})

# print()
# print(r.json())
#
# with open(r'd:\Downloads\py.html','wb') as f:
#     f.write(r.content)
#
# gpx = GPX.GPX()
# gpx_segment = GPX.GPXTrackSegment()
# gpx_track = gpx_track_create(gpx, gpx_obj_namef)
# gpx_track[gpx_obj_namef].segments.append(gpx_segment)
# gpx.description = contact_name_d
# gpx.author_email = 'andrey.korzh@atlantic.ocean.ru'
# return gpx
"""
    i_log_row_st = 0
    with OpenHDF5(cfg['in']['db_path'], cfg['in']['tables'], cfg['in']['tables_log']) as (store, tables, tables_log):
        # get last times from log or assign specified interval back from now

        # Loading
        time_intervals_dict = {}
        for tbl in cfg['in']['tables']:
            if tbl in tables:
                df_log = store.select(tables_log[tables.index(tbl)])
                _ = time_interval_default.copy()
                _['start'] = int(df_log[df_log.index[-1], 'DateEnd']).()
                time_intervals_dict[tbl] = _
            else:
                time_intervals_dict[tbl] = time_interval_default.copy()


            # mid = mid2tables[int()]
            # json = [{'mid': str(mid), **time_interval} for mid, time_interval in time_intervals_dict]
        cfg['in']['time_intervals'] = time_intervals_dict
        tbl_navs = loading(out=cfg['out'], process=cfg['process'], **cfg['in'])

        # saving(tbl_navs):

        for tbl, tbl_log in zip(tables, tables_log):
            # get last time from log
            if tbl_log:
                df_log = store.select(tbl_log,
                                      where=cfg['in']['query']
                                      )
                lf.info('Saving {} data files of ranges listed in {}', df_log.shape[0], tbl_log)

                df_log_csv = None  # order_cols(df_log, out['cols_log'])

                # df_log_csv = interp_vals(
                #     df_log_csv,
                #     df_search=None,
                #     #cols_good_data = None,
                #     db = store,
                #     dt_search_nav_tolerance = timedelta(minutes=2)
                #     )

                df_log_csv.to_csv(
                    out['text_path'] / out['file_name_fun_log'](i_log_row_st, df_log.index[0],
                                                                              df_log.DateEnd[-1]),
                    date_format=out['text_date_format'],
                    float_format=out['text_float_format'],
                    sep=out['sep']
                    )
            else:
                lf.info('{}: ', tbl)

        for i_log_row, log_row in enumerate(df_log.itertuples(),
                                            start=i_log_row_st):  # h5.log_rows_gen(table_log=tbl_log, db=store, ):
            # Load data chunk that log_row describes
            print('.', end='')
            qstr = qstr_trange_pattern.format(log_row.Index, log_row.DateEnd)
            df_raw = store.select(tbl, qstr)
            df_raw['i_log_row'] = i_log_row
            df_csv = None  # order_cols(df_raw, out['cols'])
            # Save data
            df_csv.to_csv(
                out['text_path'] / out['file_name_fun'](i_log_row, df_raw.index[0], df_raw.index[-1]),
                date_format=out['text_date_format'],
                float_format=out['text_float_format'],
                sep=out['sep']
                )

        i_log_row_st += df_log.shape[0]
"""
