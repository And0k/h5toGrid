# !/usr/bin/env python3
# coding:utf-8
"""
  Author:  Andrey Korzh <ao.korzh@gmail.com>
  Purpose: Load coordinates from GPS trackers to GPX files and HDF5 pandas store:
  - SPOT trackers data from:
    - email (from gmail.com only) messages,
    - downloaded Excel *.xlsx files;
  - Autofon trackers (http://www.autofon.ru/autofon/item/seplus) data from:
    - through http API,
    - downloaded Excel *.xlsx files.
  Created: 08.04.2021
  Modified: 08.05.2022
"""
import sys
import logging
from pathlib import Path
from shutil import copyfile
from typing import Any, Callable, Dict, Iterator, MutableMapping, Optional, List, Tuple, Union
from datetime import datetime, timedelta, timezone
from itertools import zip_longest
from dataclasses import dataclass, field
import hydra
import numpy as np
import tables
import pandas as pd
# import vaex
from pandas.tseries.frequencies import to_offset
from tables.exceptions import HDF5ExtError
import re
import requests
# import gc
from time import sleep
from tabulate import tabulate
from gpxpy.gpx import GPX
# import pyproj   # from geopy import Point, distance
# my
from h5toGpx import save_to_gpx, gpx_save  # gpx_track_create
from to_pandas_hdf5.h5_dask_pandas import df_to_csv  #, filter_global_minmax, filter_local
from to_pandas_hdf5.h5toh5 import unzip_if_need, df_log_append_fun, h5remove_tables, h5_dispenser_and_names_gen, h5load_range

import cfg_dataclasses
from utils2init import Ex_nothing_done, FakeContextIfOpen, LoggingStyleAdapter, set_field_if_no, call_with_valid_kwargs,\
    ExitStatus, GetMutex

# from csv2h5_vaex import argparser_files, with_prog_config
from to_pandas_hdf5.h5toh5 import h5move_tables, h5index_sort, h5out_init, replace_bad_db  #, h5_rem_last_rows
from to_pandas_hdf5.gpx2h5 import h5_sort_filt_append  # df_rename_cols,
from gps_tracker.mail_parse import spot_tracker_data_from_mbox, spot_from_gmail
# from inclinometer.incl_h5clc import dekart2polar_df_uv

lf = LoggingStyleAdapter(logging.getLogger(__name__))
cfg = None
tables2mid = {
    'tr0': 221910,   # MIDs of Autofon trackers are needed to load data
    'tr1': 221909,
    'tr2': 221912,
    'sp1': 4441068,
    'sp2': 2575092,  # ESNs for SPOT trackers (may be for future use)
    'sp3': 3124620,
    'sp4': 3125300,
    'sp5': 3125411,
    'sp6': 4441082
    }
mid2tables = {v: k for k, v in tables2mid.items()}

# cfg = {
#     'in': {
#         'tables': ['tr0'],  # '*',   # what to process
#         'time_intervals': {
#             'tr0': [pd.Timestamp('2021-04-08T12:00:00'), pd.Timestamp('2021-04-08T12:00:00')]},
#         'dt_from_utc': timedelta(hours=3),
#         # use already loaded coordinates instead of request:
#         'path_raw_local': {
#             'tr0': Path('d:\Work') /
#                     r' координаты адреса выходов на связь 09-04-2021 16-46-52 09-04-2021 16-46-52.xlsx'
#             },
#         'anchor': '[44.56905, 37.97308]',
#         },
#     'out': {
#         'path': Path(r'd:\WorkData\BlackSea\210408_trackers\210408trackers.h5'),
#         },
#     'process':
#         {
#             'simplify_tracks_error_m': 0,
#             'dt_per_file': timedelta(days=356),
#             'b_missed_coord_to_zeros': False
#             }
#     }


def save2gpx(nav_df: pd.DataFrame,
             track_name: str,
             path: Path = None,
             process: Dict[str, Any] = None,
             gpx: Optional[GPX] = None,
             dt_from_utc: Optional[timedelta] = None) -> GPX:
    """
    Saves track and point process['anchor_coord_default'] to the ``path / f"{nav_df.index[0]:%y%m%d_%H%M}{track_name}.gpx"``
    :param nav_df: DataFrame
    :param track_name:
    :param path: if it is
    - path of file, then save with this path and ".gpx" suffix,
    - path of directory, then autoname based on ``nav_df.index[0]`` and ``track_name`` and save in this dir,
    - None - not save, useful to get gpx only
    :param process: fields to add anchor point (will be not added if no any anchor_coord_default or anchor_coord_time):
        anchor_coord_default: List[float], [Lat, Lon] degrees, not used if anchor_coord_time specified
        anchor_coord_time: Dict[time_str, List[float]], {time_str: [Lat, Lon]} degrees. Only last item used
        anchor_depth: float, m
    :param gpx: gpxpy.gpx
    :param dt_from_utc:
    :return: updated gpx
    """
    if nav_df.empty:
        print('No data for gpx of', track_name)
        return gpx
    path_gpx = path / f'{nav_df.index[0]:%y%m%d_%H%M}{track_name}' if (path and path.is_dir()) else path
    nav_df.index.name = 'Time'
    # anchor point
    if any(process['anchor_coord_time']):
        if isinstance(a := process['anchor_coord_time'], dict):
            key_last = list(a.keys())[-1]  # only last anchor point will be in gpx (as it is most actual)
            lat_lon = process['anchor_coord_time'][key_last]
            tim = [pd.Timestamp(key_last, tz='utc')]
        else:
            # anchor track and not point. Saving
            cols = a.columns
            a = a.rename(columns={next(col := c for c in cols if c.startswith(k)): k for k in ['Lat', 'Lon']})
            tbl_anchor = col[(len('Lat') + 1):]
            gpx = save_to_gpx(
                (a if dt_from_utc is None else a.tz_convert(timezone(dt_from_utc))).dropna(subset=['Lat', 'Lon']),
                None,
                gpx_obj_namef=f'{tbl_anchor}_as_anchor', cfg_proc=process, gpx=gpx)
            lat_lon = []
    else:
        lat_lon = process['anchor_coord_default']
        if any(lat_lon) and isinstance(lat_lon[0], str):
            # we have no point of anchor coordinates (anchor position is a nav_df track or not defined)
            # todo: lat_lon = nav_df.loc[:, [f'{lat_lon[0]}Lat', f'{lat_lon[0]}Lon']].to_list()
            lat_lon = []
        tim = nav_df.index[[0]]
    # save track
    gpx = save_to_gpx(
        (nav_df if dt_from_utc is None else nav_df.tz_convert(timezone(dt_from_utc))
         ).dropna(subset=['Lat', 'Lon']),  # *.gpx will be not compatible to GPX if it will have NaN values
        None if any(lat_lon) else path_gpx,
        gpx_obj_namef=track_name, cfg_proc=process, gpx=gpx)
    if any(lat_lon):
        # save anchor
        return save_to_gpx(
            pd.DataFrame(
                {
                    'Lat': lat_lon[0],
                    'Lon': lat_lon[1],
                    'DepEcho': process['anchor_depth'],
                    'itbl': 0
                 },
                index=tim
            ),
            path_gpx,
            waypoint_symbf='Anchor',
            gpx_obj_namef= f'{track_name} mean' if process.get('b_calc_anchor_mean', False) else f'Anchor {track_name}',
            cfg_proc=process, gpx=gpx
         )
    else:
        return gpx


def prepare_loading_xlsx_links_by_pandas():
    """xlsx hyperlink support to pandas, modified original https://github.com/pandas-dev/pandas/issues/13439"""

    if prepare_loading_xlsx_links_by_pandas.is_done:
        return()

    from pandas.io.excel._openpyxl import OpenpyxlReader
    from pandas._typing import FilePathOrBuffer, Scalar
    from openpyxl.cell.cell import TYPE_BOOL, TYPE_ERROR, TYPE_NUMERIC, TIME_TYPES

    def _convert_cell(self, cell, convert_float: bool) -> Scalar:

        # here we adding this hyperlink support:
        if cell.hyperlink and cell.hyperlink.target:
            return cell.hyperlink.target
            # just for example, you able to return both value and hyperlink,
            # comment return above and uncomment return below
            # btw this may hurt you on parsing values, if symbols "|||" in value or hyperlink.
            # return f'{cell.value}|||{cell.hyperlink.target}'
        # here starts original code, except for "if" became "elif"
        else:
            cell_type = cell.data_type
            if cell_type in TIME_TYPES:
                return cell.value
            elif cell_type == TYPE_ERROR:
                return np.nan
            elif cell_type == TYPE_BOOL:
                return bool(cell.value)
            elif cell.value is None:
                return ""  # compat with xlrd
            elif cell_type == TYPE_NUMERIC:
                # GH5394
                if convert_float:
                    val = int(cell.value)
                    if val == cell.value:
                        return val
                else:
                    return float(cell.value)

        return cell.value

    def load_workbook(self, filepath_or_buffer: FilePathOrBuffer):
        from openpyxl import load_workbook
        # had to change read_only to False:
        return load_workbook(filepath_or_buffer, read_only=False, data_only=True, keep_links=False)

    OpenpyxlReader._convert_cell = _convert_cell
    OpenpyxlReader.load_workbook = load_workbook
    prepare_loading_xlsx_links_by_pandas.is_done = True


prepare_loading_xlsx_links_by_pandas.is_done = False


def autofon_df_from_dict(g, dt_from_utc: timedelta):
    """

    :param g:
    :return:

    g can contain:
     {'id': 2415919104,
      'fmt': 'CAN 0: {0} [Скорость: {2} км/ч] | CAN 1: {1} [Расход топлива: {3} л]', 'lvl': 4},  # no data
     {'id': 2550136832,
      'fmt': 'Температура 1: 0,0 С 2: 0,0 С 3: 0,0 С 4: 0,0 С', 'lvl': 5},                       # no data
     {'id': 2533359616,
      'fmt': 'Общий пробег по данным GPS/ГЛОНАСС: {0} м', 'lvl': 4},                             # no data
     {'id': 2524971008,
     'fmt': 'Курс: {0}, высота: {1}, ускорение: {2}', 'lvl': 4},                                 # rough Курс data
     {'id': 2499805184,
      'fmt': 'V1: {0} В | V2: {1} В | Vвнеш: {2} В | Vбат: {3} В', 'lvl': 4},                    # rough data
     {'id': 2155872256,
     'fmt': 'Широта: {0} Долгота: {1} Скорость: {2}', 'lvl': 4},                                 # used

     {'id': 2491416576,
      'fmt': 'Уровень сигнала GSM: {0}, HDOP: {1}, спутников GPS: {2}, спутников ГЛОНАСС: {3}, температура терминала: {4} °С',
      'lvl': 4},

     {'id': 2214789120,
      'fmt': 'Конфигурация: Параметры периода опроса координат GPS',
      'lvl': 4},
     {'id': 2214854656, 'fmt': 'Конфигурация: Первая группа параметров', 'lvl': 4},
     {'id': 2214920192, 'fmt': 'Конфигурация: Вторая группа параметров', 'lvl': 4},
     {'id': 2214985728, 'fmt': 'Конфигурация: Третья группа параметров', 'lvl': 4},
     {'id': 2215051264,
      'fmt': 'Конфигурация: Четвертая группа параметров',
      'lvl': 4},
     {'id': 2215116800, 'fmt': 'Конфигурация: Пятая группа параметров', 'lvl': 4},
     {'id': 3187671040, 'fmt': '{0}', 'lvl': 4} - на самом деле параметры
     ...
     We selected:
    id_lat_lon_speed = 2155872256
    id_lgsm_hdop_ngps_nglonass = 2491416576
    HDOP: Dilution of precision (DOP): Ideal if <= 1
    """
    id_coords = 2155872256
    id_config = 3187671040

    # used parameters
    prm = {
        id_coords: {'cols': 'Lat Lon Speed'.split(),
                    'types': np.float32},
        2491416576: {'cols': 'LGSM HDOP n_GPS n_GLONASS Temp'.split(),
                     'types': {'LGSM': np.int8, 'HDOP': np.float16, 'n_GPS': np.int8, 'Temp': np.int8},
                     'drop': ['n_GLONASS'],
                     'comma_to_dot': ['HDOP']},
        2524971008: {'cols': 'Course Height Acceleration'.split(),
                     'types': np.int8,
                     'drop': ['Height', 'Acceleration'],
                     'comma_to_dot': ['Course']},
        id_config: {'cols': 'device_settings'}
        }

    # id_lat_lon_speed = 2155872256
    # id_lgsm_hdop_ngps_nglonass = 2491416576
    # id_course_heigt_acc = 2524971008

    # cols = (['Lat', 'Lon', 'Speed'], ['LGSM', 'HDOP', 'n_GPS', 'n_GLONASS', 'Temp'])
    # cols_drop = ([], ['HDOP', 'n_GLONASS'])  # strange values for HDOP (why comma separated?), n_GLONASS always zero?
    # cols_types = (np.float32, np.int8)
    # p_lat_lon_speed = []
    # p_lgsm_hdop_ngps_nglonass = []
    p = {p_id: [] for p_id in prm}
    p_detected = {p_id: False for p_id in prm}
    for gi in g:
        try:
            (_, s_cur), (_, p_cur) = gi.items()
            assert _ == 'p'  # parameters are not interchanged
            for p_id in prm:
                if s_cur == p_id:
                    if p_detected[p_id]:
                        lf.warning('Have multiple messages for {}!', prm[p_id])  # never occurred
                        p[p_id] += p_cur
                    else:
                        p_detected[p_id] = True
                        p[p_id] = p_cur
        except:
            print(gi), print(g)
            continue

    if len(p[id_coords]) == 0:  # no data
        return ()

    if not all(p_detected.values()):
        for p_id, b_detected in p_detected.items():
            if not b_detected:
                lf.warning('Not loaded: {}!!!', prm[p_id]['cols'])
                if p_id == id_coords:  # main data absent
                    raise Ex_nothing_done

    # Show device settings
    try:
        config_last = p.pop(id_config)[-1]
        lf.info('Last config:\n{}', config_last['p'][0].replace(' |', '\n'))
    except IndexError:
        pass  # Not fatal to not display info

    del prm[id_config]  # not for output to DataFrame, no more needed
    dfs = []
    for p_id, p_cfg in prm.items():  # All parameters to ou
        p_cols = p_cfg['cols']
        # 'r' - device time, 's' - receive time, 'p' - data
        df_str = pd.DataFrame.from_records(p[p_id], index='r', exclude='s')['p']
        p_list = df_str.tolist()
        # old: drop without pandas: if 'drop' in p_cfg:
        # p_cols_keep = [c for c in p_cols if c not in p_cfg['drop']]
        # i_keep = [i for i, c in enumerate(p_cols) if c not in p_cfg['drop']]
        # p_list = [[pi[i] for i in i_keep] for pi in p_list]

        df = pd.DataFrame.from_records(
            p_list,  # to_numpy(dtype=), #.
            index=df_str.index,
            columns=p_cols,
            exclude=p_cfg.get('drop')
            )

        if 'comma_to_dot' in p_cfg:
            for c in p_cfg['comma_to_dot']:
                if p_cfg['types'] == np.int8:
                    # comma and all after is not needed:
                    df[c] = df[c].str.replace(',.*', '', regex=True)
                else:
                    df[c] = df[c].str.replace(',', '.')

        df = df.astype(p_cfg['types'], copy=False)

        df_len = len(df)
        if p_id == id_coords:  # 1st params group
            df1st = df
        else:
            dfs.append(df)
            dlen = df_len_prev - df_len
            # on merge data without pair will be ignored. Warning if more than 1 points:
            (lf.warning if dlen > 1 else lf.debug)(
                'Data lengths different, will delete {} row of ({}) > ({})',
                dlen,
                *(lambda x, y: (x, y) if dlen > 0 else (y, x))(
                    ','.join(dfs[-1].columns.tolist()), ','.join(p_cols)
                    )
                )
        df_len_prev = df_len

    # as I see indexes values of dataframes are equal so merge_asof() is not needed
    nav_df = df1st.join(dfs, how='inner')  # preferred 'left' will change int dtypes to float if dfs lengths different!
    nav_df.index = pd.to_datetime(nav_df.index, unit='s', utc=True, origin=dt_from_utc.total_seconds())
    nav_df.index.name = 'Time'

    nav_df['LGSM'] = -nav_df['LGSM']
    # if nav_df.dtypes['LGSM']
    #     nav_df['LGSM'].astype(np.uint8)
    return nav_df


def loading(
        table: str,
        path_raw_local: Union[str, Path],
        time_interval: List[pd.Timestamp],
        dt_from_utc: timedelta,
        alias
) -> pd.DataFrame:
    """
    Loads Autofon/Spot data from xlsx or Autofon server.
    Function works only for known devices listed in globdl tables2mid
    input config:
    :param table: str
    :param path_raw_local: if no then loads from Autofon server and converts to DataFrame by autofon_df_from_dict()
    :param time_interval: 2 elements list of tz-aware pandas Timestamp
    :param dt_from_utc:
    :return: pandas DataFrame
    """
    device_type, device_number_out = re.match(r'.*(sp|tr)#?(\d*).*', table).groups()
    device_number = alias.get(table, device_number_out)
    if path_raw_local:
        if path_raw_local.suffix == '.h5':
            with pd.HDFStore(path_raw_local, 'r') as db:
                nav_df = h5load_range(db, table, query_range_lims=time_interval)
            return nav_df
            # raise(Ex_nothing_done)  # not loading
        if device_type == 'sp':  # satellite based tracker
            if isinstance(path_raw_local, str):
                path_raw_local = Path(path_raw_local)
            if path_raw_local.suffix == '.xlsx':
                with pd.ExcelFile(path_raw_local) as f_xls:
                    # if f_xls.sheet_names
                    xls = pd.read_excel(f_xls,
                                        sheet_name=f'{device_number} Positions & Events',
                                        usecols='B,D',
                                        skiprows=4,
                                        index_col=0)
                    nav_df = xls['Lat/Lng'].str.extract(r'(?P<Lat>[^,]*), (?P<Lon>[^,]*)')
                nav_df.set_index((nav_df.index - dt_from_utc).tz_localize('utc'), inplace=True)
                for coord in ['Lat', 'Lon']:
                    nav_df[coord] = pd.to_numeric(nav_df[coord], downcast='float', errors='coerce')
            else:
                # Load list_of_lists from mailbox
                time_lat_lon = spot_tracker_data_from_mbox(path_raw_local,
                                subject_end=f': {device_number}',
                                time_start=time_interval[0])
                nav_df = pd.DataFrame(time_lat_lon,
                                      columns=['Time', 'Lat', 'Lon']).astype(
                                        {'Lat': np.float32, 'Lon': np.float32}
                                        )
                nav_df['Time'] -= dt_from_utc
                nav_df.set_index('Time', inplace=True)  # pd.DatetimeIndex(nav_df['Time']
                nav_df.sort_index(inplace=True)
                nav_df = nav_df.tz_localize('utc', copy=False)

        else:
            prepare_loading_xlsx_links_by_pandas()
            xls = pd.read_excel(path_raw_local, usecols='C:D', skiprows=4, index_col='Дата')
            nav_df = xls['Адрес / Координаты'].str.extract(r'[^=]+\=(?P<Lat>[^,]*),(?P<Lon>[^,]*)').astype(
                                        {'Lat': np.float32, 'Lon': np.float32}
                                        )
            nav_df.set_index((nav_df.index - dt_from_utc).tz_localize('utc'), inplace=True)
            # for coord in ['Lat', 'Lon']:
            #     nav_df[coord] = pd.to_numeric(nav_df[coord], downcast='float', errors='coerce')

        nav_df = nav_df.truncate(*time_interval, copy=False)  # [k] - dt_from_utc for k in [0,1] .tz_localize('utc')
        tim_last_coord = pd.Timestamp.now(tz='utc')
    elif device_type == 'sp':  # satellite based tracker
        # download and parse from GMail
        time_lat_lon = spot_from_gmail(
            device_number=device_number,
            time_start=time_interval[0] - dt_from_utc)
        nav_df = pd.DataFrame(time_lat_lon,
                              columns=['Time', 'Lat', 'Lon']).astype(
            {'Lat': np.float32, 'Lon': np.float32}
            )
        nav_df['Time'] -= dt_from_utc
        nav_df.set_index('Time', inplace=True)  # pd.DatetimeIndex(nav_df['Time']
        nav_df.sort_index(inplace=True)
        nav_df = nav_df.tz_localize('utc', copy=False)
    else:
        mid = tables2mid[table]
        url = 'http://176.9.114.139:9002/jsonapi'
        key_pwd = 'key=d7f1c7a5e53f48a5b1cb0cf2247d93b6&pwd=ao.korzh@yandex.ru'

        # Request and display last info
        try:
            r = requests.post(f'{url}/laststates/?{key_pwd}')
        except TimeoutError:
            print('TimeoutError - skip')
            raise (Ex_nothing_done)
        if r.status_code != 200:
            print(r)
            raise(Ex_nothing_done)
        for d in r.json():
            if d['id'] != mid:
                continue
            tim_last_coord = pd.Timestamp(datetime.fromtimestamp(d['tscrd']), tz=timezone(dt_from_utc))
            lf.info(f'Last coordinates on server for #{{id}}: {tim_last_coord}, (N{{lat}}, E{{lng}})'.format_map(d))

        # Request and save new data

        if False:  # this will obtain filtered data useful for display only
            r = requests.post(f'{url}/?{key_pwd}',
                              json=[{'mid': str(k), **time_interval} for k in tables2mid.keys()]
                              )
            if r.status_code != 200:
                print(r)
                raise(Ex_nothing_done)
            nav_df = pd.DataFrame.from_records(d['points'], index='ts') \
                .rename(columns={'lat': 'Lat', 'lng': 'Lon', 'v': 'Speed'})

        r = requests.post(
            f"{url}/messages/{mid}?{key_pwd}&fromdate={{}}&todate={{}}&minlevel=6".format(
                *[int(t.timestamp()) for t in time_interval]))  # &lang=en not works
        if r.status_code != 200:
            print(r)
            raise(Ex_nothing_done)
        nav_df = autofon_df_from_dict(r.json()[0]['items'], dt_from_utc)

    if len(nav_df) == 0:
        lf.info(f"{len(nav_df)} points got")
        # have all data already => return without error: if no new data and time from last download is less than minute
        if time_interval[1] - tim_last_coord.tz_convert('utc') > timedelta(minutes=1):
            lf.warning(f"No data interval: {time_interval[1] - tim_last_coord.tz_convert('utc')}")
        raise(Ex_nothing_done)

    # display last point with local time
    lf.info(f"{table}: got {len(nav_df)} points, last - "
            f"{nav_df.index[-1].tz_convert(timezone(dt_from_utc))}: "
            "{Lat:0.6f}N, {Lon:0.6f}E".format_map(nav_df.iloc[-1]))
    return nav_df


def saving(nav_dfs, path, process):
    for tbl, nav_df in nav_dfs.items():
        save2gpx(nav_df, tbl, path=path, process=process)

        # bin average data
        bins = ['1H', '5min']
        for bin in bins:
            save2gpx(nav_df.resample(bin).mean(), f'{tbl}_avg({bin})', path=path, process=process)


def proc(cfg=cfg):
    cur_time = datetime.now()
    time_interval = {
        'start': int((cur_time - timedelta(days=100)).timestamp()),
        # 1st data to load, will be corrected if already have some
        'end': int(cur_time.timestamp())
        }

    nav_dfs = []
    for tbl in cfg['in']['tables']:
        nav_dfs[tbl] = call_with_valid_kwargs(loading,
                                              time_interval=time_interval,
                                              **cfg['in']
                                              )
    saving(nav_dfs,
           path=cfg['out']['path'].parent,
           process=cfg['process']
           )


class OpenHDF5(FakeContextIfOpen):
    """
    Context manager that do nothing if file is not str/PurePath or custom open function is None/False
    useful if instead file want use already opened file object
    Returns table names and opened store

    :param tables: tables names search pattern or sequence of table names
    :param tables_log: tables names for metadata of data in `tables`
    :param db_path:
    :return: iterator that returns (table name, coefficients)
    updates cfg_in['tables'] - sets to list of found tables in store
    """
    # will be filled by each table from cfg['in']['tables']

    def __init__(self, db_path, tables, tables_log, db=None):
        self.tables = tables
        self.tables_log = tables_log
        self.handle = super(OpenHDF5, self).__init__(
            lambda f: pd.HDFStore(f, mode='r'),
            file=db_path,
            opened_file_object=db)

    def __enter__(self):
        """
        :return: opened handle or :param file: from __init__ if not need open
        """
        self.handle = super(OpenHDF5, self).__enter__()
        return (
            self.handle,
            self.tables,
            self.tables_log or [f'{t}/log' for t in self.tables]
            )

hydra.output_subdir = 'cfg'
# hydra.conf.HydraConf.output_subdir = 'cfg'
# hydra.conf.HydraConf.run.dir = './outputs/${now:%Y-%m-%d}_${now:%H-%M-%S}'


@dataclass
class ConfigInAutofon:
    time_interval: List[str] = field(default_factory=lambda: ['2021-04-08T12:00:00', 'now'])  # UTC
    # data coordinates source (path_raw_local - single, path_raw_local - multiple) with value of file path or None:
    # - None: request data from internet,
    # - xls/xlsx-file: load its data instead,
    # - h5-file: loading and use its data for reprocess (same as if table is included in tbl_raw_not_update below):
    path_raw_local_default: Optional[str] = None
    # If there are different devices' data in different sources then use dict {device: source} where device is regex
    # pattern str. that matches devices name that is output table name defined by ``out`` config. If no key will be
    # found then path_raw_local_default value will be used
    path_raw_local: Optional[Dict[str, str]] = field(default_factory=dict)
    dt_from_utc_hours: int = 0
    # b_incremental_update: bool = True
    tbl_raw_not_update: List[str] = field(default_factory=lambda: [])  # List tables in .raw.h5 that not try to update
    alias: Optional[Dict[Any, str]] = field(default_factory=dict)  # load device with name=value  and output with
    # name=key which corresponds to output table

@dataclass
class ConfigProcessAutofon:
    # gpx track settings
    simplify_tracks_error_m = 0
    dt_per_file_days = 356
    b_missed_coord_to_zeros: bool = False
    period_tracks: Optional[str] = None
    period_segments: Optional[str] = '1D'
    # anchor settings
    anchor_coord_default: Any = field(default_factory=lambda: [])    # List[float]: constant coord. (i.e. [44.56905, 37.97308]),
    anchor_coord: Dict[Any, Any] = field(default_factory=dict)  # {tracker: [Lat, Lon]} - anchor for each tracker
    # or str "mean" to assign mean of each data source to anchor coord else - empty list to use anchor_coord_time
    anchor_coord_time: Dict[Any, Any] = field(default_factory=dict)  # {time: [Lat, Lon]} - use if anchor moved
    # If anchor_coord_default and anchor_coord_time are empty, and no tracker key in anchor_coord too then will not
    # calc. distance.
    anchor_depth: float = 0
    anchor_tracker: List[str] = field(default_factory=lambda: [])   # names of devices to calc. distances to them as to anchors
    max_dr_default: float = 100  # common maximum distance to anchor, m. Delete data with dr > max_dr
    max_dr: Dict[str, float] = field(default_factory=dict)  # maximum distance to anchor for specified output tables

    # absent data settings
    # detect absent data to try download again. Default is '10min' for gprs and '20min' for satellite based tracker:
    dt_max_hole: Optional[str] = None
    # detect absent data only in this interval back from last data. Default is '1D' for GPRS and '0D' for satellite based tracker:
    dt_max_wait: Optional[str] = None

    # other
    # b_reprocess: bool = False  # todo: after implementing of processing only last data part after its loading: will be useful to reprocess all data (previously calculated dx, dy...) from Lat, Lon, DateTimeIndex
    interp_limit: int = 5    # 120 * 5min/60 = 10h


@dataclass
class ConfigOutAutofon(cfg_dataclasses.ConfigOutSimple):
    # dt_bins_rolling: List[List[str]] = field(default_factory=lambda: [['2H', None], ['5min', None], ['10min', '1H']])
    # List[List[ or List[Optional[str] not supported so we split it:
    dt_bins: List[str] = field(default_factory=lambda: ['2H', '5min', '10min'])
    dt_rollings: List[str] = field(default_factory=lambda: ['', '', '1H'])
    # len(to_gpx) should be less than 1(for raw table) + dt_bins, if empty then output tho gpx if averaging less than 1h.
    to_gpx: List[bool] = field(default_factory=list)


ConfigProgram = cfg_dataclasses.ConfigProgram

cs_store_name = Path(__file__).stem
cs, ConfigType = cfg_dataclasses.hydra_cfg_store(
    cs_store_name, {
    'input': [ConfigInAutofon],  # Load the config "in_autofon" from the config group "input"
    'out': [ConfigOutAutofon],  # Set as MISSING to require the user to specify a value on the command line.
    #'filter': ['filter'],
    'process': [ConfigProcessAutofon],  # 'process_autofon'
    'program': [ConfigProgram],  # 'program'
    # 'search_path': 'empty.yml' not works
    },
    module=sys.modules[__name__]
    )


def dx_dy_dist_bearing(lon1, lat1, lon2, lat2):
    """
    Distance and bearing between two points
    :param lon1: degrees, 1st point "lon"-coordinate
    :param lat1: degrees, 1st point "lat"-coordinate
    :param lon2: degrees, 2nd point "lon"-coordinate
    :param lat2: degrees, 2nd point "lat"-coordinate
    :return: array with 4 columns:
     - dx: m, distance along "lon" coord line
     - dy: m, distance along "lat" coord line
     - dist: m, distance between points
     - bearing: degrees, in the range ``[-180, 180]``
    """
    lon1, lat1, lon2, lat2 = map(np.radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1

    R = 6371000  # radius of the earth in m
    klon = np.cos((lat2+lat1)/2)
    dx = R * klon * dlon
    dy = R * dlat

    d = np.sqrt(dy**2 + dx**2)
    # angle = np.arctan2(dlat, dlon)
    angle = np.arctan2(dx, dy)  # or use atan2[(sin Δλ ⋅ cos φ₂), (cos φ₁ ⋅ sin φ₂ − sin φ₁ ⋅ cos φ₂ ⋅ cos Δλ)] from https://www.omnicalculator.com/other/azimuth

    return np.column_stack((dx, dy, d, np.degrees(angle)))

    #haversine_
    # x = sin(dlon) * cos(lat2)
    # y = cos(lat1) * sin(lat2) - (sin(lat1) * cos(lat2) * cos(diffLong))
    #
    #
    # m = 6367000 * 2 * np.arcsin(np.sqrt(a))
    #
    # dx = np.cos(lat1) * np.cos(lat2) * np.sin(dlon/2)
    # dy =np.sin(dlon / 2.0)
    # a = np.sin(dlat / 2.0)


def format_log_filename(start, end):
    return '{:%y%m%d_%H%M}-{:%m%d_%H%M}'.format(start, end)


def proc_and_h5save(df, tbl, cfg_in, out, process: MutableMapping[str, Any],
                    bin: Optional[str] = None, rolling_dt: Optional[str] = None):
    """
    Calculates displacement, bin average and saves to HDF5
    For averaged data tables if column Course exist then calculates averaged 'speed_x' and 'speed_y' from Course and dr
     with dropping 'Course'.
    :param df: DataFrame with datetime index and columns:
        Lat, Lon: coordinates
    :param tbl: table name where to save result in Pandas HDF5 store
    Configuration dicts:
    :param cfg_in:
    :param out: dict, output config
    :param process: dict, processing config with fields:
    - b_calc_anchor_mean: if True then calc anchor coordinates by averaging data and update process['anchor_coord_default']
    - anchor_coord_time: either
      - Dict where keys are times and values are position of anchor. process['anchor_coord_default'] is used before 1st item.
      - Dataframe with position of anchor, which will be interpolated with regular interval (bin_raw=5min)
    - other fields initialised from ConfigProcessAutofon
    Averaging parameters:
    :param bin:
    :param rolling_dt:
    :return:
    Modifies:
     - df: columns of polar data and what will be calculated after average are removed,
     - out['table'] = tbl + suffix "ref{tbl_anchor}" if anchor is DataFrame with coord. columns "Lat_{tbl_anchor}",
        "Lon_{tbl_anchor}"
     - out['tables_written'] in h5_sort_filt_append()
     - process['anchor_coord_default']
    """

    if 'n_GPS' in df.columns:
        min_n_GPS = 3
        b_bad = df.n_GPS < min_n_GPS
        b_bad_sum = b_bad.sum()
        if b_bad_sum:
            lf.warning('{} rows with number of GPS satellites < {} deleted!', b_bad_sum, min_n_GPS)
            df.drop(df[b_bad].index, inplace=True)

    if df.empty:
        return ()

    out['log']['fileChangeTime'] = datetime.now(tz=timezone.utc)  # :%y%m%d_%H%M%S
    out['log']['fileName'] = format_log_filename(*df.index[[0, -1]])
    try:
        del out['log']['index']  # was temporary used internally
    except KeyError:
        pass                     # may be deleted already

    if bin is not None:
        # Drop Vdir and Course because it is not correct to average angles: will be calculated after average along with
        # other columns: ['dx', 'dy', 'dr', 'Vdir'] from <Lat>, <Lon>.
        # 'Course' we need to be convert to dekart: For vector length we using `dr` instead `Speed` which is too rough
        # dekart2polar_df_uv(df.rename())
        cols_need_recalc = ('dr', 'Vdir')
        if {'Course', 'dr'}.issubset(df.columns):  # (if no anchor specified then no dr because not need proc)
            # if already done then no 'dx' because of side effects here, same as 'Course' (to not repeat or try if no Course)
            # reuse 2 float columns instead of dropping, then drop other:
            df.rename(columns={'dx': 'speed_x', 'dy': 'speed_y'}, inplace=True)
            course = np.radians(df['Course'].values)

            df['speed_x'], df['speed_y'] = df['dr'].values * np.vstack([np.sin(course), np.cos(course)])
            df = df.drop(['dr', 'Vdir', 'Course'], axis='columns')

        # Remove cols we will recalc that should not be averaged directly
        if any(cols_dropped := [c for c in df.columns if c.startswith(cols_need_recalc)]):
            df = df.drop(cols_dropped, axis='columns')

        # bin average data
        shift_to_mid = to_offset(bin) / 2
        df = df.resample(bin, offset=-shift_to_mid).mean()
        df.index += shift_to_mid

        if rolling_dt:
            window_not_even = int(re.match('.*mov(\d+)bin', tbl).group(1))
            lf.info('{}: moving average over {} bins of {}', tbl, window_not_even, bin)
            out['log']['fileName'] += f'bin{bin}'
            df = df.resample(bin, origin='epoch').mean().rolling(
                window_not_even, center=True, win_type='gaussian', min_periods=1).mean(std=3)
            out['log']['fileName'] += f'avg{rolling_dt}'
        else:
            lf.info('{}: {}-bin average', tbl, bin)
            out['log']['fileName'] += f'avg{bin}'

    if process['b_calc_anchor_mean']:
        if not (isinstance(process['b_calc_anchor_mean'], str) and
                tbl.startswith(process['b_calc_anchor_mean'])):
          # use anchor_coord_default from original table if it is suffixed
            process['b_calc_anchor_mean'] = tbl  # save to not recalc for suffixed tables
            process['anchor_coord_default'] = df.loc[:, ['Lat', 'Lon']].values.mean(axis=0).tolist()  # seems numpy mean result in better accuracy
            lf.info('set anchor position to mean {} position: {}', tbl, process['anchor_coord_default'])
    anchor_coord = process['anchor_coord_default']

    tbl_anchor = ''
    if any(a := process['anchor_coord_time']):
        if isinstance(a, dict):
            anchor_lat_lon = np.zeros((2, len(df)), np.float32)
            if any(anchor_coord):
                anchor_lat_lon += np.float32(anchor_coord)[:, None]
            anchor_times = pd.DatetimeIndex(a.keys(), tz='utc')
            i_starts = np.searchsorted(df.index, anchor_times).tolist()
            for i_st, i_en, lat_lon in zip(i_starts, i_starts[1:] + [len(df)], a.values()):
                anchor_lat_lon[:, i_st:i_en] = np.float32(lat_lon)[:, None]
            anchor_coord = anchor_lat_lon[::-1, :]
        else:  # dataframe with pairs of cols Lat*, Lon*. todo: There may be multiple pairs
            cols = [next(col := c for c in a.columns if c.startswith(k)) for k in ['Lon', 'Lat']]
            if cols[0] not in df.columns:
                tbl_anchor = col[(len('Lat') + 1):]
                if tbl == tbl_anchor:  # this device is used as anchor for other device
                    anchor_coord = []
                else:
                    bin_raw = '5min'
                    lf.info('interpolating {} and {} as moving anchor on regular {}-bin intervals to calc '
                            'distanses between',
                            tbl, tbl_anchor, bin_raw
                            )
                    new_index = pd.date_range(
                        min(df.index[0], a.index[0]).floor(bin_raw),
                        max(df.index[-1], a.index[-1]).ceil(bin_raw),
                        freq=bin_raw
                        )
                    df_upsample = df.reindex(df.index.union(new_index)).interpolate(
                        'index', limit=process['interp_limit']).reindex(new_index)  # method='spline', order=2
                    index_union = a.index.union(new_index)
                    try:
                        a_upsample = a.reindex(index_union)
                    except ValueError:  # cannot reindex from a duplicate axis
                        if any(dups := index_union.duplicated()):
                            a_upsample = a.reindex(index_union[~index_union.duplicated()])
                            lf.warning('{} duplicates on input!', dups.sum())
                    a_upsample = a_upsample.interpolate('index', limit=process['interp_limit']).reindex(new_index)

                    df = pd.concat([df_upsample, a_upsample], axis=1)  # sort=False,
                    # a.reindex(df.index).interpolate(method='krogh')
                    anchor_coord = a_upsample[cols].values.T
                    # df.drop(cols, axis=1, inplace=True)
    else:
        anchor_coord = anchor_coord[::-1]  # Lat, Lon -> Lon, Lat

    # Calculate parameters
    out['table'] = f"{tbl}_ref_{tbl_anchor}" if tbl_anchor else tbl
    if len(anchor_coord):
        df.loc[:, ['dx', 'dy', 'dr', 'Vdir']] = dx_dy_dist_bearing(
            *anchor_coord,
            *df[['Lon', 'Lat']].values.T
            )
        # course, azimuth2, distance = geod.inv(  # compute forward and back azimuths, plus distance
        #     *df.loc[navp_d['indexs'][[0, -1]], ['Lon', 'Lat']].values.flat)  # lon0, lat0, lon1, lat1

        # Filter source data
        if bin is None and (max_dr := process['max_dr'].get(out['table'], process['max_dr_default'])):
            b_bad = df.dr > max_dr
            b_bad_sum = b_bad.sum()
            if b_bad_sum:
                if process['b_calc_anchor_mean']:
                    lf.warning('{} rows with dr > {} found!', b_bad_sum, max_dr)
                    process['anchor_coord_default'] = df[~b_bad].loc[:, ['Lat', 'Lon']].mean().tolist()
                    lf.info('updated anchor coord: {}', process['anchor_coord_default'])
                    # recalc r and b_bad to filter with updated anchor coord
                    anchor_coord = process['anchor_coord_default'][::-1]  # Lat, Lon -> Lon, Lat
                    df.loc[:, ['dx', 'dy', 'dr', 'Vdir']] = dx_dy_dist_bearing(
                        *anchor_coord,
                        *df[['Lon', 'Lat']].values.T
                        )
                    b_bad = df.dr > max_dr
                    b_bad_sum = b_bad.sum()
                    if b_bad_sum:
                        lf.warning('{} rows with dr > {} deleted!', b_bad_sum, max_dr)
                        df.drop(df[b_bad].index, inplace=True)
                else:
                    lf.warning('{} rows with dr > {} deleted!', b_bad_sum, max_dr)
                    df.drop(df[b_bad].index, inplace=True)

    # Recalc cols_dropped
    if bin is not None and any(cols_dropped):
        sfx_processed = []
        for c in cols_dropped:
            for cc in cols_need_recalc:
                cc += '_'
                if len(cc) < len(c) and c.startswith(cc):  # 1st excludes empty suffixes
                    sfx = c[len(cc):]
                    if sfx in sfx_processed:
                        continue
                    sfx_processed.append(sfx)
                    df = df.assign(**{
                        f'dr_{sfx}': np.sqrt(df[f'dx_{sfx}']**2 + df[f'dy_{sfx}']**2),
                        f'Vdir_{sfx}': np.arctan2(df[f'dx_{sfx}'], df[f'dy_{sfx}']) * (180 / np.pi)
                        })
                    break

    # Saving to HDF5
    return h5_sort_filt_append(df, input={**cfg_in, 'dt_from_utc': timedelta(0)}, out=out)


def holes_starts(t: pd.DatetimeIndex, t_max: int) -> Tuple[pd.DatetimeIndex, np.timedelta64]:
    """
    Finds time starts of data holes and its sizes
    :param t: time data
    :param t_max: max time difference considered not to be hole
    :return: t_start, dt
    """
    dt_int = np.ediff1d(t.view(np.int64), to_end=0)  # or .to_numpy(dtype=np.int64)
    i_hole = np.flatnonzero(dt_int > t_max)
    dt = np.array(dt_int[i_hole], 'm8[ns]').astype('m8[s]')  # np.round(dt_int[i_hole]*1E-9, 1)
    t_start = t[i_hole]
    n_holes = len(t_start)
    if n_holes:
        n_show = 20
        msg_number = f'Last {n_show} of {n_holes}' if n_holes > n_show else f'Found {n_holes}'
        lf.warning('{} holes:\n{}', msg_number,
            tabulate({'time_start': t_start[-n_show:],
                      'dt_minutes': dt.astype(int)[-n_show:] / 60},
                     headers='keys',
                     floatfmt='.1f'
                     )
                   )
    else:
        lf.debug('no holes found')
    return t_start, dt


def holes_prepare_to_fill(db, tbl, tbl_log,
                          time_holes: Optional[List[pd.Timestamp]] = None,
                          dt_max_hole: Optional[str] = '10min',
                          dt_max_wait: Optional[str] = '1D'
                          ) -> Tuple[Optional[List[pd.Timestamp]], Optional[pd.Timestamp], Optional[str]]:
    """
    Finds time holes in data index (if time_holes is None else just uses it instead) to download from data's hole start

    :param db: opened pandas HDF5 data store
    :param tbl: table name
    :param tbl_log: log table name
    Search holes parameters:
    :param time_holes:  None to search holes. If DatetimeIndex then return it as time data gaps.  Will select 1st gap > max_time - dt_max_wait
    :param dt_max_hole: Search holes of this size at least
    :param dt_max_wait: Search/select holes only in this range from end:
        None: search everywhere/not change
        '0': do not use but search for 1 day and display message
    :return:
     time_holes: list of found holes' starts
     time_start: time of last data if no holes we want to fill else hole start (starting Timestamp to query server)
     msg: log info of reason for time_start used: 'last hole' or 'last saved'
    """

    try:  # Last data time from log table
        t_max_exist = db.select(tbl_log, columns=['DateEnd'], start=-1)['DateEnd'][0].tz_convert(tz=timezone.utc)
    except (KeyError, IndexError, AttributeError) as e:      # no log (yet or lost?)
        try:
            t_min_exist = db.select(tbl, columns=[], stop=1).index[0]
            t_max_exist = db.select(tbl, columns=[], start=-1).index[0]
            lf.warning('{}! Appending one row for all data: {} - {}',
                       e.msg if hasattr(e, 'msg') else str(e), t_min_exist, t_max_exist
                       )
            # try:
            #     t_date_end = t_max_exist.tz_convert(db.select(tbl_log, columns=['DateEnd']).dtypes['DateEnd'].tz)
            # except:
            #     lf.exception('old format converting fail but may be not needed if it is new')
            #     t_date_end = t_max_exist  # to try just as is
            try:
                n_rows = db.get_storer(tbl).group.table.shape[0]
            except:
                n_rows = -1
            df_log = pd.DataFrame(
                {'DateEnd': [t_max_exist],
                 'fileChangeTime': datetime.now(tz=timezone.utc),
                 'fileName': format_log_filename(t_min_exist, t_max_exist),
                 'rows': n_rows},
                index=[t_min_exist]
            )
            df_log_append_fun(df_log, tbl_log, {'db': db, 'nfiles': None})
        except (KeyError, IndexError, AttributeError):  # no data yet  # not need if all worked properly
            return time_holes, None, 'start'


    time_start_wait = t_max_exist - to_offset(dt_max_wait).delta if dt_max_wait else None
    max_hole_timedelta = to_offset(dt_max_hole)

    if time_holes is None:
        # searching holes
        for try_query in (
                        [f"index > '{time_start_wait}'", ''] if time_start_wait else
                        [''] if time_start_wait is None else
                        ['1D']):
            try:
                time_holes, dt_holes = holes_starts(db.select(tbl, try_query, columns=[]).index,
                                                    max_hole_timedelta.nanos)  # nanoseconds
                break
            except Exception as e:  # tables.exceptions.NoSuchNodeError  # have only once, may be not needed
                if try_query:       # 'UnImplemented' object has no attribute 'description'
                    lf.error('Can not query {} for "{}": {}. Retrying in memory', tbl, try_query, e)
                    continue  # Checking for holes all data
                else:
                    raise HDF5ExtError('Can not load data from "{}/{}"'.format(db, tbl))
    elif time_start_wait:
        time_holes = time_holes[time_holes >= time_start_wait]

    if len(time_holes) and time_start_wait:
        msg_start_origin = 'last hole'
        t_start = time_holes[0]

        # try:
        #     df_log_cur = db.select(
        #         tbl_log,
        #         f"index <= Timestamp('{t_start}') & "
        #         f"DateEnd > Timestamp('{t_start}')")
        # except Exception as e:  # tables.exceptions.NoSuchNodeError  # have only once, may be not needed
        #     lf.error('Can not query {} for "{}": {}. Doing in memory', tbl_log, try_query, e)
        #     df_log_cur = db[tbl_log]
        #     b_good = (df_log_cur.index <= t_start) & (t_start < df_log_cur.DateEnd)
        #     df_log_cur = df_log_cur[b_good]
        #
        # # log_dict['fileName'] =
        # # log_dict['fileChangeTime'] = df_log_cur['fileChangeTime'].iat[0]
        #
        # if len(df_log_cur) != 1:                # duplicates, how?
        #     df_log_cur = df_log_cur.iloc[[0]]   # to del. next rows = duplicates
        # if not h5_rem_last_rows(db, [tbl, tbl_log], [df_log_cur], t_start):
        #     return time_holes, None, 'new start'            # failed to use previous data
    else:
        msg_start_origin = 'last saved'
        t_start = t_max_exist
    return time_holes, t_start, msg_start_origin

    # df_log[df_log['fileName'] == log['fileName']]
    # df_log = cfg_out['db'][tbl_log]
    # t_max_exist = df_log['DateEnd'].max()
    #
    #
    # qstr = "index>=Timestamp('{}')".format(df_log_cur.index[0])
    # qstrL = "fileName=='{}'".format(df_log_cur['fileName'][0])
    #
    # # df = cfg_out['db'][tbl]
    # df = cfg_out['db'].select(tbl, where=qstr)


def h5_names_gen(cfg_in: MutableMapping[str, Any],
                 cfg_out: MutableMapping[str, Any],
                 processed_tables: bool,
                 dt_max_hole: Optional[str] = None,
                 dt_max_wait: Optional[str] = None,
                 **kwargs) -> Iterator[Union[Tuple[str, str, Optional[str], str, bool],
                                             Tuple[int, str, str, str, Optional[str], str, bool]]]:
    """
    Generate table names from cfg_out['tables'] and other parameters depending on ``processed_tables`` mode parameter,
    sets cfg_out['log'] fields 'fileChangeTime' and 'fileName' depending on cfg_in['time_interval'],

    :param cfg_in: fields:
      - 'time_interval': changes here for each talble to query time interval depending on existed data
    :param cfg_out:
      - 'time_interval': required output data time interval (initially defined and copied here from ConfigInAutofon)
    :param processed_tables: True => function will work in mode of generating parameters to average already loaded data:
    generates 1st raw table then processed table names according to averaging as defined by 'dt_bins' and 'dt_rollings'.
    :param dt_max_hole: holes_prepare_to_fill() parameter:
    :param dt_max_wait: holes_prepare_to_fill() parameter:
    :param kwargs:
    # time_holes: holes_prepare_to_fill() parameter: list of found holes' starts
    # time_start: holes_prepare_to_fill() parameter: starting Timestamp to query server
    :return: (tbl, tbl_log, t_start, msg_start_origin, to_gpx) or
             (tbl, tbl_log, bin, rolling_dt, to_gpx):
    (`t_start`, msg_start_origin) replaced with (`bin`, rolling_dt) if processed_tables
    t_start, msg_start_origin: see holes_prepare_to_fill()

    Modifies: cfg_in['time_interval'], cfg_out['log']
    """
    if not processed_tables:
        loading_str = 'loading' if cfg_in['path_raw_local_default'] else 'downloading'
        msg_start_fmt = '{} {} {:%y-%m-%d %H:%M:%S} \u2013 {:%m-%d %H:%M:%S UTC}{}'

    set_field_if_no(cfg_out, 'log', {})
    cfg_in['new_data_time_starts'] = {}  # save here existed data ends to know later from where new data starts

    # tables_log_copy = cfg_out['tables_log']
    for itbl, tbl in enumerate(cfg_out['tables'], start=1):
        cfg_out['log']['fileChangeTime'] = cfg_out['time_interval'][-1]
        cfg_out['log']['fileName'] = format_log_filename(*cfg_out['time_interval'])
        tbl_log = cfg_out['tables_log'][0].format(tbl)
        # cfg_out['tables_log'] = [tbl_log]
        try:
            if not processed_tables:
                # mode 1: check existed data for holes and output its last good time: t_start
                lf.info(f'{tbl} -------------------- Checking existed data...')
                tbl_raw_not_update_copy = cfg_in['path_raw_local_default']  # to update cfg_in before yield and then revert
                for retry in [False, True]:
                    try:
                        time_holes, t_start, msg_start_origin = holes_prepare_to_fill(cfg_out['db'], tbl, tbl_log,
                                                                          time_holes=kwargs.get('time_holes', None),
                                                                          dt_max_hole=dt_max_hole,
                                                                          dt_max_wait=dt_max_wait)
                    except HDF5ExtError:
                        lf.exception('Bad DB table. Recovering table...')
                        try:
                            h5remove_tables(cfg_out['db'], [tbl], [])
                            cfg_out['db'].close()
                        except:
                            lf.exception('Bad DB. Recovering table failed. Recovering DB.'
                                         'Deleting temporary to replacing by output store...')
                            cfg_out['db'].close()
                            cfg_out['db_path_temp'].unlink()
                        try:
                            h5move_tables(cfg_out, tbl_names=[tbl])
                        except KeyError:
                            lf.exception('no data?')
                            t_start = None
                            msg_start_origin = 'beginning again'
                            retry = False
                        cfg_out['db'] = pd.HDFStore(cfg_out['db_path_temp'])
                        if retry:
                            lf.exception('Bad DB. Deleting failed')
                            raise
                        else:
                            continue

                # Skipping updating source table or custom raw source conditions
                b_skip = False
                if tbl in cfg_in['tbl_raw_not_update']:
                    str_skip = 'Skipping loading device (as configured)'
                    b_skip = True  # to  skip to next tbl after source checking message
                for k, v in cfg_in['path_raw_local'].items():
                    if re.match(k, tbl):
                        cfg_in['path_raw_local_default'] = v    # will be used as loading() argument
                        break

                # Del. data after t_start in h5del_obsolete() called after this fun. by h5_dispenser_and_names_gen():
                if cfg_in['path_raw_local_default'] != cfg_out['db_path']:  # but not if raw source == saving source file
                    cfg_out['log']['index'] = t_start + timedelta(microseconds=1) if t_start else None
                    if not b_skip:
                        str_skip = f'Continue {loading_str}'  # i.e. not skip
                else:
                    b_skip = True  # to skip to next tbl after source checking message
                    str_skip = f'{loading_str.capitalize()} existed data in store only'

                # Set ``time_interval`` argument for loading()
                cfg_in['time_interval'] = cfg_out['time_interval'].copy()  # set to required interval, then reduce:
                if t_start is None:  # previous data last time not found - Ok, maybe we have no data yet
                    lf.info(msg_start_fmt, f'{str_skip}, but no data found!' if b_skip else loading_str.capitalize(),
                            tbl, *cfg_in['time_interval'], '...')
                else:
                    cfg_in['time_interval'][0] = t_start
                    if msg_start_origin == 'last saved':
                        lf.info(msg_start_fmt, str_skip, tbl, *cfg_in['time_interval'],
                                f' (from {msg_start_origin})')
                        cfg_in['time_interval'][0] += timedelta(
                            seconds=1)  # else 1st downloaded row will be same as the last we have
                    elif t_start:  # 'last hole'
                        lf.warning(msg_start_fmt, str_skip, tbl, *cfg_in['time_interval'], ': filling gaps. Note:' +
                                   f"{cfg_out['time_interval'][-1] - t_start} from {msg_start_origin} to now!")
                    else:  # possible?
                        lf.error('t_start = {}!', t_start)
                cfg_in['new_data_time_starts'][tbl] = cfg_in['time_interval'][0]

                if b_skip:
                    continue
                yield (tbl, tbl_log, None, msg_start_origin, cfg_out['to_gpx'][0])
                cfg_in['path_raw_local_default'] = tbl_raw_not_update_copy
                # cfg_out['tables_log'] = tables_log_copy
            else:
                # mode 2: Generate parameters for tables that will be processed
                lf.info(f'{tbl}: processing')
                # todo: load last data time - averaging interval shift, check last data for each table
                t_start = cfg_in.get('t_good_last')
                for bin, rolling_dt, to_gpx in zip_longest(
                        cfg_out['dt_bins'],
                        cfg_out['dt_rollings'],
                        cfg_out['to_gpx'][1:]
                        ):
                    # Table name form averaging parameters
                    bin_timedelta = to_offset(bin)
                    if rolling_dt:
                        rolling_timedelta = to_offset(rolling_dt)
                        window_not_even = int(rolling_timedelta / bin_timedelta)
                        if not (window_not_even % 2):
                            window_not_even += 1
                        tbl_avg = f'{tbl}_avg{rolling_dt.lower()}=mov{window_not_even}bin{bin.lower()}'
                        max_timedelta = max(bin_timedelta, rolling_timedelta)
                    else:
                        tbl_avg = f'{tbl}_avg{bin.lower()}'
                        max_timedelta = bin_timedelta

                    # default not output to gpx if averaging is bigger than 1 hours
                    if to_gpx is None:
                        to_gpx = max_timedelta <= timedelta(hours=1)
                    cfg_out['log']['index'] = t_start  # will be read in h5del_obsolete() called after this fun. to del.
                    tbl_log = cfg_out['tables_log'][0].format(tbl_avg)
                    cfg_tables_save, cfg_out['tables'] = cfg_out['tables'], [tbl_avg]  # for compatibility with del_obsolete()
                    yield (itbl, tbl_avg, tbl_log, bin, rolling_dt, to_gpx)
                    cfg_out[tables] = cfg_tables_save                                  # recover
        except GeneratorExit:
            print('Something wrong?')
            return ExitStatus.failure


def h5move_and_sort(out: MutableMapping[str, Any]):
    """
    Moves from temporary storage and sorts `tables_written` tables and clears this list
    :param out: fields:
        tables_written: set of tuples of str, table names
        b_del_temp_db and other fields from h5index_sort
    :return:
    Modifies out fields:
        b_remove_duplicates: True
        tables_written: assigns to empty set

    """

    failed_storages = h5move_tables(out, tbl_names=out['tables_written'])
    print('Finishing...' if failed_storages else 'Ok.', end=' ')
    # Sort if have any processed data, else don't because ``ptprepack`` not closes hdf5 source if it not finds data
    out['b_remove_duplicates'] = True
    h5index_sort(out,
                 out_storage_name=f"{out['db_path'].stem}-resorted.h5",
                 in_storages=failed_storages,
                 tables=out['tables_written']
                 )
    out['tables_written'] = set()


@hydra.main(config_name=cs_store_name, config_path="cfg", version_base='1.3')  # adds config store cs_store_name data/structure to :param config data/structure to :param config
def main(config: ConfigType) -> None:
    """
    ----------------------------
    Save data to GPX files and Pandas HDF5 store*.h5
    ----------------------------
    The store contains tables for each device and each device table contains log with metadata of recording sessions

    :param config: with fields:
    - in: mapping with fields:
      - tables_log: - log table name or pattern str for it: in pattern '{}' will be replaced by data table name

    - out: mapping with fields:

    """
    if GetMutex().IsRunning():
        lf.info("Application is already running. Waiting 10s...")
        sleep(10)
        sys.exit(ExitStatus.failure)

    global cfg
    cfg = cfg_dataclasses.main_init(config, cs_store_name)
    cfg_in = cfg.pop('input')
    cfg_in['cfgFile'] = cs_store_name
    cfg['in'] = cfg_in
    # back to string keys because can't suppress hydra str to int conversion:
    cfg['in']['alias'] = {str(k): v for k, v in cfg['in']['alias'].items()}
    # try:
    #     cfg = to_vaex_hdf5.cfg_dataclasses.main_init_input_file(cfg, cs_store_name, )
    # except Ex_nothing_done:
    #     pass  # existed db is not mandatory
    # geod = pyproj.Geod(ellps='WGS84')
    # dir_create_if_need(out['text_path'])
    out = cfg['out']
    # if relative path for cfg['out']['db_path'] then it is from directory of running script
    if cfg['out'].get('db_path') and not cfg['out']['db_path'].is_absolute():
        cfg['out']['db_path'] = Path(sys.argv[0]).parent / cfg['out']['db_path']
    out_db_path = cfg['out']['db_path'].with_suffix('.h5')
    out['raw_db_path'] = cfg['out']['db_path'] = out_db_path.with_suffix('.raw.h5')

    h5out_init(cfg['in'], out)

    if not out['to_gpx']:  # default to output to gpx raw data and for averaged only if averaging is less than 1 hours
        out['to_gpx'] = [True]  # output 1st, other (will be set to None) depends on averaging
    out['tables_written'] = set()

    # Device specific default parameters to check existed data
    b_sp = all('sp' in t for t in out['tables'])  # Spot trackers => min data period is big (>=5min) and no memory
    if cfg['process']['dt_max_hole'] is None:
        cfg['process']['dt_max_hole'] = '20min' if b_sp else '10min'
    if cfg['process']['dt_max_wait'] is None:
        cfg['process']['dt_max_wait'] = '0D' if b_sp else '1D'

    # out['b_incremental_update'] = True  # (default) # enables to update only resent data using info of last data we have
    out['field_to_del_older_records'] = 'index'   # new data priority (based on time only)
    out['b_reuse_temporary_tables'] = True  # do not copy result tables to temp one by one instead use result store copy
    for t in cfg_in['time_interval']:
        print(t.tzinfo)
    out['time_interval'] = [
        t.replace(tzinfo=timezone.utc) if ~(isinstance(t, str) and t == 'now') else
        datetime.now(tz=timezone.utc) for t in cfg_in['time_interval']
        ]
    cfg_in['new_data_time_starts'] = {}  # default, will be updated in h5_names_gen with last data time in store
    
    # flag to assign anchor position as its mean coordinates in proc_and_h5save() for each device:
    cfg['process']['b_calc_anchor_mean'] =\
        isinstance(cfg['process']['anchor_coord_default'], str) and cfg['process']['anchor_coord_default'] == 'mean'
    
    lf.debug('updating raw data...')
    df_loaded = None
    for b_retry in [False, True]:  # if updating fails then will recover db from *.copy.h5 and try again
        # Loading data for each probe to temporary store cycle
        ######################################################
        for i1_tbl, (tbl, tbl_log, _, msg_start_origin, _) in h5_dispenser_and_names_gen(
                cfg_in, out,
                fun_gen=h5_names_gen,  # here we prepare loading
                processed_tables=False,
                dt_max_hole=cfg['process']['dt_max_hole'],
                dt_max_wait=cfg['process']['dt_max_wait']
                ):

            lf.info('{}. {}: ', i1_tbl, tbl)
            # Loading (file or data from Internet)
            try:
                df_loaded = call_with_valid_kwargs(
                    loading, table=tbl,
                    **{**cfg_in, 'path_raw_local': cfg_in['path_raw_local'].get(tbl, cfg_in['path_raw_local_default'])}
                )
            except (Ex_nothing_done, TimeoutError):
                if i1_tbl < len(out['tables']):  # need to check other devices
                    continue
                elif not (df_loaded is None or out_db_path.is_file()):  # need to process
                    break                        # go to processing
                lf.info('No new data. Exiting...')
                sys.exit(ExitStatus.failure)

            # Update new received/loaded data to HDF5
            # todo: Save current data last good time and tables for which it is valid to reduce searching good start on next run
            # df_loaded.index[-1]
            del out['log'][out['field_to_del_older_records']]  # remove field temporary used to del. overwritten data
            h5_sort_filt_append(df_loaded, input={**cfg_in, 'dt_from_utc': timedelta(0)},
                                out={**out, 'table': f'{tbl}'})


        # Recover temp db if were errors
        if out.get('db_is_bad') and not b_retry:
            lf.warning(f'Recovering temp db from *.copy.h5 because of "db_is_bad" flag have been set')
            try:
                copyfile(src=out['db_path'].with_suffix('.copy.h5'), dst=out['db_path_temp'])
                out['db_is_bad'] = False
                continue
            except:
                lf.exception(f'Recovering temp db from *.copy.h5 failed!')
        sleep(2)  # reduces error rate if all bad
        break

    # Temporary to result store
    if any(unzip_if_need(out['tables_written'])):
        # todo: same for raw imput h5-data if from different sourcev
        # cfg_in.get('time_last'):
        try:
            if out.get('db_is_bad'):
                db_path_temp_copy = replace_bad_db(out['db_path_temp'], out['db_path'])
                lf.warning('Trying copy {} from {} file to {}', out['tables'], db_path_temp_copy, out['db_path_temp'])
                h5move_and_sort(
                    {'db_path': out['db_path_temp'],
                     'db_path_temp': db_path_temp_copy,
                     'tables_written': out['tables']
                     })
                sleep(2)
            h5move_and_sort(out)
        except HDF5ExtError:
            if out.get('db_is_bad'):
                lf.exception('Temp store is bad. Nothing helps to delete. Delete it manually!')
            else:
                replace_bad_db(out['db_path_temp'], out['db_path'])
            sys.exit(ExitStatus.failure)

    # Update reserved and temp store with copy of (compressed, sorted and indexed) result store:
    try:
        if not out.get('db_is_bad'):  # update reserved only if were no errors
            lf.info('saving copy to *.copy.h5')
            copyfile(src=out['db_path'], dst=out['db_path'].with_suffix('.copy.h5'))
        copyfile(out['db_path'], out['db_path_temp'])
    except FileNotFoundError:
        lf.exception('Have no data yet')
        sys.exit(ExitStatus.failure)
    except:
        lf.exception('Can not replace temporary file {}', out['db_path_temp'])

    # Process data
    ##############

    out['b_reuse_temporary_tables'] = False
    out['db_path'] = out_db_path
    out['db_path_temp'] = out_db_path.with_suffix('.not_sorted.h5')
    lf.info('Processing loaded data to {}...', cfg['out']['db_path'])
    #print('\tSaving to GPX and other processing')
    # time start to reprocess data with account for averaging period  # todo: use it to reprocess only needed data part
    time_end_st = min(cfg_in['new_data_time_starts'].values()) - max(
        to_offset(str_dt) for str_dt in out['dt_bins'] + out['dt_rollings'] if str_dt
        )
    gpx_raw = None
    gpx = {}  # navigation for different averaging
    df_anch = None  # will keep coordinates of cfg['process']['anchor_tracker'] if it is not empty
    tbl_raw_prev = None   # for tbl.startswith(tbl_raw) return False at first
    tbls_no_anch_tr = []
    # tables counter
    tbls_set = set(cfg_in['new_data_time_starts'].keys())
    n_tables = len(cfg['process']['anchor_tracker'] + list(
        tbls_set.difference(cfg['process']['anchor_tracker'])
    ))
    qstr_range_pattern = "index>='{}' & index<='{}'"
    qstr = qstr_range_pattern.format(*out['time_interval'])
    for i1_tbl, (i1_tbl_raw, tbl, tbl_log, bin, rolling_dt, b_to_gpx) in h5_dispenser_and_names_gen(
        cfg_in, out,
        fun_gen=h5_names_gen,
        processed_tables=True
        # dt_max_hole=cfg['process']['dt_max_hole'],  # dt_max_wait=cfg['process']['dt_max_wait']
        ):

        # load data needed to reprocess
        tbl_raw, sfx = tbl.split('_', 1)
        if tbl_raw != tbl_raw_prev:  # next raw table processing
            tbl_raw_prev = tbl_raw

            lf.info('{}. Averaging {} with {}: ', i1_tbl_raw, tbl_raw, sfx)
            with pd.HDFStore(out['raw_db_path'], mode='r') as store_raw:
                # Load concatenated current and previous raw data
                df_raw = store_raw.select(tbl_raw, where=qstr)

                # 1. Raw data relative to anchor with parameters relative to constant anchor cfg['process']['anchor_coord_default'] or do nothing
                df_no_avg = proc_and_h5save(df_raw, tbl_raw, cfg_in, out, cfg['process'])

                # 2. Interpolated coord. and parameters relative to device defined by cfg['process']['anchor_tracker']
                if tbl_raw in cfg['process']['anchor_tracker']:
                    assert i1_tbl_raw == 1
                    # remember
                    df_anch = df_raw.rename(columns={col: f'{col}_{tbl_raw}' for col in df_raw.columns})
                    tbl_anch = tbl_raw
                elif df_anch is not None:  # process and save: interp, update with params relative to anchor, with df_anch if it is not None
                    df_no_avg = proc_and_h5save(
                        df_raw, tbl_raw, cfg_in, out,
                        {**cfg['process'], 'anchor_coord_time': df_anch}
                        )
                    # Mark params relative to anchor (to keep and save after averaging)
                    # if df_anch is not None:
                    df_no_avg.rename(inplace=True, columns={
                        col: f'{col}_{tbl_raw}_ref_{tbl_anch}' for col in df_raw.columns if col.startswith(
                            ('dx', 'dy', 'dr', 'Vdir'))
                        })

                # Saving GPX: accumulate raw data in gpx_raw and save on last cycle
                if out['to_gpx'][0]:
                    gpx_raw = save2gpx(
                        df_raw,
                        track_name=tbl_raw,
                        path=(out['db_path'].with_name(
                                    f"{df_raw.index[0]:%y%m%d_%H%M}{','.join(sorted(tbls_set))}"
                                ) if i1_tbl_raw == n_tables else None  # accumulates only if None
                              ),
                        process=cfg['process'],
                        dt_from_utc=cfg_in['dt_from_utc'],
                        gpx=gpx_raw
                        )
        else:
            lf.info('Averaging {} with {}: ', tbl_raw, sfx)

        # Average tracks calc., save to other tables in HDF5
        # ##################################################

        df = proc_and_h5save(
            df_no_avg,  # if tbl_raw == tbl_loaded else out['db'].select(tbl_raw),
            tbl, cfg_in, out, cfg['process'], bin=bin, rolling_dt=rolling_dt
            )

        # Saving averaged GPX
        if b_to_gpx:
            # Anchor(s) gpx track to memory - if have corresponded columns in df (suffixed by device id)
            tbls = cfg['process']['anchor_tracker'] if df_anch is not None and f'Lat_{tbl_anch}' in df.columns else []
            if tbls:
                if sfx in gpx and tbls_no_anch_tr:  # have accumulated data
                    # # save it (not tested because was no case to save with and without anchor tracker)
                    # gpx_save(gpx[sfx], gpx_obj_namef=out['db_path'].stem, cfg_proc=cfg['process'], path_stem=(
                    #     out['db_path'].with_name(
                    #         f"{df_raw.index[0]:%y%m%d_%H%M}{','.join(tbls_no_anch_tr)}"
                    #     )))
                    
                    # Start accumulate for new anchor tracker file
                    del gpx[sfx]
                    tbls_no_anch_tr = []

                for tbl_anch in tbls:
                    cols_anch = {f'{c}_{tbl_anch}': c for c in ['Lat', 'Lon']}
                    gpx[sfx] = save2gpx(
                        df[cols_anch.keys()].rename(columns=cols_anch), f'{tbl_anch}_{sfx}',
                        path=None, process=cfg['process'],
                        dt_from_utc=cfg_in['dt_from_utc'],
                        gpx=gpx.get(sfx, None)
                        )
            else:
                # to accumulate gpx for data without anchor tracker
                tbls_no_anch_tr.append(tbl_raw)

            # Accumulate for different suffixes and write gpx at last tbl_raw cycle
            gpx[sfx] = save2gpx(
                df[['Lat', 'Lon']], tbl,
                path=(out['db_path'].with_name(
                    f"{df_raw.index[0]:%y%m%d_%H%M}{','.join([tbl_raw] + tbls)}{'_' if tbls else ''}{sfx}"
                    ) if i1_tbl_raw == n_tables else None
                ),
                process=cfg['process'],
                dt_from_utc=cfg_in['dt_from_utc'],
                gpx=gpx.get(sfx, None)
                )

        # Saving CSV
        df_to_csv(df, {
            'db_path': out['db_path'],
            'table': tbl
            })

    if any(unzip_if_need(out['tables_written'])):
        # not need to remove temp db if want use out['b_reuse_temporary_tables'] option: but temporary tables remains else set out['b_del_temp_db'] = True
        h5move_and_sort(out)

    try:
        lf.debug('Update processed temp store with copy of (compressed, sorted and indexed) result store')
        copyfile(out['db_path'], out['db_path_temp'])
    except:
        lf.exception('Can not replace temporary file {}', out['db_path_temp'])
    print('Ok>', end=' ')


def load_prev_source_data(tbl: str, tbl_log: str, store: pd.HDFStore, db_path_other: Path, select_where=''):
    """
    Try load from df, df_log from tbl, tbl_log tables of store or from db_path_other
    :param tbl:
    :param tbl_log:
    :param select_where:
    :param store:
    :param db_path_other:
    :return: df, df_log

    """
    df = None
    df_log = None
    try:  # normally have data in opened temp store:
        df = store.select(tbl, where=select_where)
        df_log = store.select(tbl_log, where=select_where)
    except KeyError:  # 'No object named tr2 in the file': Was table occasionally deleted?
        lf.info('Nothing found to reprocess in temp db')
    if df is None or df_log is None or df_log.empty:
        try:  # if failed then load from output store:
            with pd.HDFStore(db_path_other, mode='r') as store_in:
                df = store_in.select(tbl, where=select_where)
                df_log = store_in.select(tbl_log, where=select_where)
        except OSError:  # ... .h5` does not exist
            lf.info('Nothing to reprocess in db. Not need load/remove')
        except KeyError:  # ... .h5` does not exist
            lf.error('Not found {} in {}? Nothing to reprocess in db. Can not load/remove previous data. Continue...',
                     tbl, db_path_other
                     )
    return df, df_log


def main_call(
        cmd_line_list: Optional[List[str]] = None,
        fun: Callable[[], Any] = main
        ) -> Dict:
    """
    Adds command line args, calls fun, then restores command line args
    :param cmd_line_list: command line args of hydra commands or config options selecting/overwriting

    :return: global cfg
    """

    sys_argv_save = sys.argv.copy()
    if cmd_line_list is not None:
        sys.argv += cmd_line_list

    # hydra.conf.HydraConf.run.dir = './outputs/${now:%Y-%m-%d}_${now:%H-%M-%S}'
    fun()
    sys.argv = sys_argv_save
    return cfg


if __name__ == '__main__':
    main()  # [f'--config-dir={Path(__file__).parent}'])


def pairwise(iterable):
    """s -> (s0, s1), (s2, s3), (s4, s5), ..."""
    a = iter(iterable)
    return zip(a, a)


def call_example():
    """
    to run from IDE or from bat-file with parameters
    --- bat file ---
    call conda.bat activate py3.7x64h5togrid
    D: && cd D:/Work/_Python3/And0K/h5toGrid
    python -c "from to_vaex_hdf5.autofon_coord import call_example; call_example()"
    ----------------
    # python -m to_vaex_hdf5.autofon_coord.call_example() not works
    :return:
    """
    # from to_vaex_hdf5.h5tocsv import main_call as h5tocsv
    path_db = Path(
        r'd:/workData/BalticSea/210515_tracker/current@tr0/210515_1500tr0.h5'
        # r'd:\WorkData\BlackSea\210408_trackers\tr0\210408trackers.h5'
        )
    device = ['tr0']  # 221912
    main_call([  # '='.join(k,v) for k,v in pairwise([   # ["2021-04-08T08:35:00", "2021-04-14T11:45:00"]'
        'input.time_interval=[2021-05-15T13:00:00, now]',  # ["2021-04-08T09:00:00", "now"]',   # UTC, max (will be loaded and updated what is absent)
        'input.dt_from_utc_hours=2',  #3
        'process.anchor_coord_default=[54.62425, 19.76050]', #[44.56905, 37.97309]',
        'process.anchor_depth=20',
        'process.period_tracks=1D',
        # 'process.period_segments="2H"', todo: not implemented to work if period_tracks is set
        f'out.db_path="{path_db}"',
        'out.tables=[{}]'.format(','.join([f'"{d}"' for d in device])),
        # 'out.tables_log=["{}/log"]',
        # 'out.b_insert_separator=False'
        # 'input.path_raw_local_default="{}"'.format({  # use already loaded coordinates instead of request:
        #     221910: Path('d:\Work') /
        #             r' координаты адреса выходов на связь 09-04-2021 08-03-13 09-04-2021 08-03-13.xlsx'
        #     }),
        # 'input.time_intervals={221910: "2021-04-08T12:00:00"}',

        ])

    # conda activate py3.7x64h5togrid && D: && cd D:\Work\_Python3\And0K\h5toGrid && python -m to_vaex_hdf5.autofon_coord.call_example()

# r = requests.post(
#     'http://176.9.114.139/jsonapi/?key=5cee5887daa94dff88e51dc9bd6d16a2&pwd=0887219675',  # &a0ghzybxtr
#     json={'mid': 'ao', **time_interval})

# print()
# print(r.json())
#
# with open(r'd:\Downloads\py.html','wb') as f:
#     f.write(r.content)
#
# gpx = GPX.GPX()
# gpx_segment = GPX.GPXTrackSegment()
# gpx_track = gpx_track_create(gpx, gpx_obj_namef)
# gpx_track[gpx_obj_namef].segments.append(gpx_segment)
# gpx.description = contact_name_d
# gpx.author_email = 'andrey.korzh@atlantic.ocean.ru'
# return gpx
"""
    i_log_st = 0
    with OpenHDF5(cfg['in']['db_path'], cfg['in']['tables'], cfg['in']['tables_log']) as (store, tables, tables_log):
        # get last times from log or assign specified interval back from now

        # Loading
        time_intervals = {}
        for tbl in cfg['in']['tables']:
            if tbl in tables:
                df_log = store.select(tables_log[tables.index(tbl)])
                _ = time_interval_default.copy()
                _['start'] = int(df_log[df_log.index[-1], 'DateEnd']).timestamp()
                time_intervals[tbl] = _
            else:
                time_intervals[tbl] = time_interval_default.copy()


            # mid = mid2tables[int()]
            # json = [{'mid': str(mid), **time_interval} for mid, time_interval in time_intervals]
        cfg['in']['time_intervals'] = time_intervals
        tbl_navs = loading(out=cfg['out'], process=cfg['process'], **cfg['in'])

        # saving(tbl_navs):

        for tbl, tbl_log in zip(tables, tables_log):
            # get last time from log
            if tbl_log:
                df_log = store.select(tbl_log,
                                      where=cfg['in']['query']
                                      )
                lf.info('Saving {} data files of ranges listed in {}', df_log.shape[0], tbl_log)

                df_log_csv = None  # order_cols(df_log, out['cols_log'])

                # df_log_csv = interp_vals(
                #     df_log_csv,
                #     df_search=None,
                #     #cols_good_data = None,
                #     db = store,
                #     dt_search_nav_tolerance = timedelta(minutes=2)
                #     )

                df_log_csv.to_csv(
                    out['text_path'] / out['file_name_fun_log'](i_log_st, df_log.index[0],
                                                                              df_log.DateEnd[-1]),
                    date_format=out['text_date_format'],
                    float_format=out['text_float_format'],
                    sep=out['sep']
                    )
            else:
                lf.info('{}: ', tbl)

        for i_log, log_row in enumerate(df_log.itertuples(),
                                            start=i_log_st):  # h5log_rows_gen(table_log=tbl_log, db=store, ):
            # Load data chunk that log_row describes
            print('.', end='')
            qstr = qstr_trange_pattern.format(log_row.Index, log_row.DateEnd)
            df_raw = store.select(tbl, qstr)
            df_raw['i_log'] = i_log
            df_csv = None  # order_cols(df_raw, out['cols'])
            # Save data
            df_csv.to_csv(
                out['text_path'] / out['file_name_fun'](i_log, df_raw.index[0], df_raw.index[-1]),
                date_format=out['text_date_format'],
                float_format=out['text_float_format'],
                sep=out['sep']
                )

        i_log_st += df_log.shape[0]
        


    # Loading data for each probe to temporary store cycle
    ######################################################
    for i1_tbl, (tbl, tbl_log, t_start, msg_start_origin, _) in h5_dispenser_and_names_gen(
            cfg_in, out,
            fun_gen=h5_names_gen,
            processed_tables=False,
            dt_max_hole=cfg['process']['dt_max_hole'],
            dt_max_wait=cfg['process']['dt_max_wait']
            ):
        lf.info('{}. {}: ', i1_tbl, tbl)
        df_loaded = []


        ########## loading
        df_loaded = out['db'][tbl]
        del out['log'][out['field_to_del_older_records']]  # remove field temprary used to del. overwritten data
        h5_sort_filt_append(
            df_loaded, input={**cfg_in, 'table': f'{tbl}_raw', 'dt_from_utc': timedelta(0)}, out=out
            )

        # Reprocess option: load previous source data, append new data, delete previous table to be ready to write.
        if cfg['process']['b_reprocess']:
            lf.info('prepend all previous stored source data in memory, and reprocess')
            df, df_log = load_prev_source_data(
                tbl, tbl_log, out['db'], out['db_path'], select_where=f"index < '{df_loaded.index[-1]}'")
            if df is not None:  # selected df.index < df_loaded.index[-1]
                df_loaded = df[df_loaded.columns].append(df_loaded)
                h5remove_tables(out['db'], [tbl], [])
                if df_log.empty:  # empty log is not normal but skip if so
                    lf.warning('log is empty')
                else:
                    # with ReplaceTableKeepingChilds([df], tbl, cfg, df_log_append_fun):
                    #     pass
                    df_log_append_fun(df_log, tbl_log, {**out, **{'nfiles': None}})
                    # h5_append({'table': tbl, **out}, [], df_log, log_dt_from_utc=cfg_in['dt_from_utc'])









    # Loading concatenated current and previous data. todo: only needed time range
    b_all_needed_data_in_temp_store = False
    with pd.HDFStore(out['db_path'], mode='r') as store_raw:
        df_raw = store_raw.select(tbl)  # we just saved to output (indexed) store so loading it from there
        tbl_loaded = tbl
        # copy all tables in order keeping last used table last to continue processing on it below loop
        if not cfg_in.get('t_good_last'):
            with tables.open_file(out['db_path_temp'], mode='w') as store_out:
                for tbl in ([t for t in store_raw.root.__members__ if t != tbl] + [tbl]):
                    # Overwriting temp store

                    # will process all data => not need older processed tables. Removing them by
                    # replace temporary store with raw table (compressed, sorted) without indexes from result store:
                    try:
                        store_raw._handle.copy_node(
                            f'/{tbl}',
                            newparent=store_out.root,
                            recursive=True,
                            overwrite=True)
                    except AttributeError:  # AttributeError: 'numpy.int64' object has no attribute '_pack'
                        lf.exception('Overwriting temp store error.')
                    b_all_needed_data_in_temp_store = True
                store_out.flush()  # .flush(fsync=True
                # out_back = out
                # out_back['db_path'], out_back['db_path_temp'] = out_back['db_path_temp'], out_back['db_path']
                # out_back['db_path']
                # h5move_and_sort({'db_path': out[], 'tables_written': tables_written})


    # # remove raw data from output store if not need:
    # if b_all_needed_data_in_temp_store:
    #     # out['tables_written'] = tables_written
        #     out['db_path'].unlink()  # remove file
        
        
"""
